import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


def create_reports_folder(session_folder: Path) -> Path:
    """Create and return the reports directory for the current session."""
    reports_folder = session_folder / "reports"
    reports_folder.mkdir(parents=True, exist_ok=True)
    return reports_folder


def save_csv_report(
    session_folder: Path,
    test_data: dict[str, Any],
) -> Path:
    """Save the current test result as a CSV file."""
    reports_folder = create_reports_folder(session_folder)
    report_path = reports_folder / "test_result.csv"

    with report_path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(
            csv_file,
            fieldnames=list(test_data.keys()),
        )

        writer.writeheader()
        writer.writerow(test_data)

    return report_path


def save_text_report(
    session_folder: Path,
    test_data: dict[str, Any],
) -> Path:
    """Save a readable text summary of the current test."""
    reports_folder = create_reports_folder(session_folder)
    report_path = reports_folder / "test_summary.txt"

    with report_path.open("w", encoding="utf-8") as text_file:
        text_file.write("WNC TESTHUB - TITAN 3 TEST SUMMARY\n")
        text_file.write("=" * 45 + "\n\n")

        for key, value in test_data.items():
            readable_key = key.replace("_", " ").title()

            if value is None or value == "":
                displayed_value = "Not provided"
            else:
                displayed_value = value

            text_file.write(
                f"{readable_key}: {displayed_value}\n"
            )

    return report_path


def save_json_report(
    session_folder: Path,
    test_data: dict[str, Any],
) -> Path:
    """Save structured test data as JSON."""
    reports_folder = create_reports_folder(session_folder)
    report_path = reports_folder / "test_result.json"

    with report_path.open("w", encoding="utf-8") as json_file:
        json.dump(
            test_data,
            json_file,
            indent=4,
            default=str,
        )

    return report_path


def save_excel_report(
    session_folder: Path,
    test_data: dict[str, Any],
) -> Path:
    """
    Append the current test result to an Excel workbook.

    The first test creates the workbook and header row.
    Every test after that is appended as a new row.
    """
    reports_folder = create_reports_folder(session_folder)
    report_path = reports_folder / "test_results.xlsx"

    if report_path.exists():
        workbook = load_workbook(report_path)
        worksheet = workbook["Test Results"]
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Test Results"

        headers = list(test_data.keys())
        worksheet.append(headers)

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        worksheet.freeze_panes = "A2"

    existing_headers = [
        cell.value
        for cell in worksheet[1]
    ]

    new_headers = [
        key
        for key in test_data.keys()
        if key not in existing_headers
    ]

    if new_headers:
        for header in new_headers:
            worksheet.cell(
                row=1,
                column=worksheet.max_column + 1,
                value=header,
            )

        existing_headers.extend(new_headers)

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

    row_values = [
        test_data.get(header, "")
        for header in existing_headers
    ]

    worksheet.append(row_values)

    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        maximum_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                maximum_length = max(
                    maximum_length,
                    len(str(cell.value)),
                )

        worksheet.column_dimensions[column_letter].width = min(
            maximum_length + 2,
            40,
        )

    workbook.save(report_path)

    return report_path


def generate_reports(
    session_folder: Path,
    test_data: dict[str, Any],
) -> list[Path]:
    """Generate all currently supported report formats."""
    return [
        save_csv_report(session_folder, test_data),
        save_text_report(session_folder, test_data),
        save_json_report(session_folder, test_data),
        save_excel_report(session_folder, test_data),
    ]