from __future__ import annotations

from datetime import datetime
from pathlib import Path
from io import BytesIO
import csv
import json
import os
import re
import shutil
import subprocess
import webbrowser
from threading import Lock
from typing import Any
from uuid import uuid4

from fastapi import BackgroundTasks, FastAPI, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from openpyxl import Workbook, load_workbook
import requests
from PIL import Image
import pytesseract

from automation.automated_runner import AutomatedTestRunner
from automation.throughput_test import ThroughputTester
from automation.verizon_syslog_automation import (
    automate_verizon_syslog_download,
)
from config import (
    DEFAULT_TITAN_IP,
    QXDM_DEFAULT_LOG_FILENAME,
    QXDM_DEFAULT_MASK,
    QXDM_EXECUTABLE,
    QXDM_MAX_LOG_SIZE_MB,
    RESULTS_FOLDER,
)
from controllers.qxdm_controller import QXDMController
from controllers.syslog_controller import SyslogController
from wrapper.test_wrapper import TestWrapper
from titan3 import Titan3
from utils import (
    create_session_folder,
    create_session_folders,
    create_session_record,
    find_session_record,
    list_session_records,
)


app = FastAPI(
    title="WNC TestHub API",
    version="1.0.0",
)

# Allow the Vite React frontend to call this backend.
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ==========================================================
# Throughput models and state
# ==========================================================

class ThroughputRequest(BaseModel):
    titan_ip: str = Field(
        default=DEFAULT_TITAN_IP,
        min_length=7,
        max_length=45,
    )
    number_of_runs: int = Field(
        default=5,
        ge=1,
        le=15,
    )
    delay_between_runs: int = Field(
        default=10,
        ge=0,
        le=3600,
    )
    timeout_seconds: int = Field(
        default=180,
        ge=1,
        le=1800,
    )


class ThroughputJob(BaseModel):
    job_id: str
    status: str
    message: str
    titan_ip: str
    number_of_runs: int
    completed_runs: int = 0
    results: list[dict[str, Any]] = Field(default_factory=list)
    error: str | None = None
    session_folder: str | None = None
    excel_path: str | None = None


class ThroughputGUILaunchRequest(BaseModel):
    titan_ip: str = Field(
        default=DEFAULT_TITAN_IP,
        min_length=7,
        max_length=45,
    )


class ThroughputGUIResultRequest(BaseModel):
    job_id: str = Field(
        min_length=1,
        max_length=100,
    )
    download_mbps: float = Field(ge=0)
    upload_mbps: float = Field(ge=0)
    ping_ms: float = Field(ge=0)
    ping_jitter_ms: float | None = Field(
        default=None,
        ge=0,
    )
    packet_loss_percent: float | None = Field(
        default=None,
        ge=0,
        le=100,
    )
    server_name: str = Field(
        min_length=1,
        max_length=200,
    )
    server_location: str = Field(
        default="",
        max_length=200,
    )
    server_id: str | None = Field(
        default=None,
        max_length=100,
    )
    isp: str | None = Field(
        default=None,
        max_length=200,
    )
    result_url: str | None = Field(
        default=None,
        max_length=1000,
    )
    notes: str = Field(
        default="",
        max_length=2000,
    )


class SpeedtestResultLinkRequest(BaseModel):
    result_url: str = Field(
        min_length=10,
        max_length=1000,
    )


class ThroughputCSVImportRequest(BaseModel):
    job_id: str = Field(
        min_length=1,
        max_length=100,
    )
    wrapper_session_folder: str | None = Field(
        default=None,
        max_length=500,
    )


class SessionCreateRequest(BaseModel):
    session_name: str = Field(
        min_length=1,
        max_length=120,
    )
    titan_ip: str = Field(
        default=DEFAULT_TITAN_IP,
        min_length=7,
        max_length=45,
    )
    notes: str = Field(
        default="",
        max_length=2000,
    )


class TestSession(BaseModel):
    session_id: str
    session_name: str
    titan_ip: str
    notes: str = ""
    status: str
    created_at: str
    updated_at: str
    session_folder: str
    throughput_jobs: list[dict[str, Any]] = Field(
        default_factory=list
    )
    qxdm_logs: list[dict[str, Any]] = Field(
        default_factory=list
    )
    reports: list[dict[str, Any]] = Field(
        default_factory=list
    )


class AnalyticsSummary(BaseModel):
    total_runs: int = 0
    average_download_mbps: float | None = None
    minimum_download_mbps: float | None = None
    maximum_download_mbps: float | None = None
    average_upload_mbps: float | None = None
    minimum_upload_mbps: float | None = None
    maximum_upload_mbps: float | None = None
    average_ping_ms: float | None = None
    average_jitter_ms: float | None = None


class AnalyticsResponse(BaseModel):
    summary: AnalyticsSummary
    history: list[dict[str, Any]] = Field(default_factory=list)


jobs: dict[str, dict[str, Any]] = {}
jobs_lock = Lock()


def update_job(
    job_id: str,
    **changes: Any,
) -> None:
    with jobs_lock:
        current_job = jobs.get(job_id)

        if current_job is None:
            return

        current_job.update(changes)



def build_speedtest_png_url(
    result_url: str,
) -> str:
    """
    Convert a normal Speedtest result URL into the static PNG result URL.
    """
    cleaned = result_url.strip()

    if not re.match(
        r"^https://(?:www\.)?speedtest\.net/my-result/",
        cleaned,
        flags=re.IGNORECASE,
    ):
        raise ValueError(
            "Enter a valid Speedtest result URL beginning with "
            "https://www.speedtest.net/my-result/."
        )

    if cleaned.lower().endswith(".png"):
        return cleaned

    return f"{cleaned}.png"


def _extract_metric_value(
    text: str,
    label: str,
) -> float | None:
    """
    Extract a numeric value appearing after a label in OCR text.
    """
    patterns = [
        rf"{label}\s*[:\-]?\s*([0-9]+(?:\.[0-9]+)?)",
        rf"{label}[^0-9]{{0,40}}([0-9]+(?:\.[0-9]+)?)",
    ]

    for pattern in patterns:
        match = re.search(
            pattern,
            text,
            flags=re.IGNORECASE,
        )

        if match:
            try:
                return float(match.group(1))
            except (TypeError, ValueError):
                continue

    return None


def parse_speedtest_png_text(
    text: str,
    result_url: str,
    png_url: str,
) -> dict[str, Any]:
    """
    Parse OCR text from a Speedtest PNG into TestHub throughput fields.
    """
    normalized_text = " ".join(
        text.replace("\r", "\n").split()
    )

    result: dict[str, Any] = {
        "result_url": result_url,
        "png_url": png_url,
        "download_mbps": _extract_metric_value(
            normalized_text,
            "download",
        ),
        "upload_mbps": _extract_metric_value(
            normalized_text,
            "upload",
        ),
        "ping_ms": _extract_metric_value(
            normalized_text,
            "ping",
        ),
        "ping_jitter_ms": _extract_metric_value(
            normalized_text,
            "jitter",
        ),
        "packet_loss_percent": _extract_metric_value(
            normalized_text,
            "loss",
        ),
        "server_name": None,
        "server_location": None,
        "ocr_text": normalized_text,
    }

    server_patterns = [
        r"Server\s*[:\-]?\s*([A-Za-z0-9 .&'_-]{2,80})",
        r"Hosted by\s+([A-Za-z0-9 .&'_-]{2,80})",
    ]

    for pattern in server_patterns:
        match = re.search(
            pattern,
            normalized_text,
            flags=re.IGNORECASE,
        )
        if match:
            value = match.group(1).strip()
            value = re.split(
                r"\b(?:download|upload|ping|jitter|loss)\b",
                value,
                maxsplit=1,
                flags=re.IGNORECASE,
            )[0].strip()
            if value:
                result["server_name"] = value
                break

    location_match = re.search(
        r"([A-Za-z .'-]+,\s*[A-Z]{2})\b",
        normalized_text,
    )

    if location_match:
        result["server_location"] = (
            location_match.group(1).strip()
        )

    return result


def fetch_speedtest_png_result(
    result_url: str,
) -> dict[str, Any]:
    """
    Download the Speedtest PNG and OCR the visible result values.
    """
    png_url = build_speedtest_png_url(
        result_url
    )

    try:
        response = requests.get(
            png_url,
            timeout=20,
            allow_redirects=True,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/151.0 Safari/537.36"
                ),
                "Accept": (
                    "image/png,image/*;q=0.9,*/*;q=0.8"
                ),
            },
        )
        response.raise_for_status()

    except requests.RequestException as error:
        raise RuntimeError(
            "TestHub could not download the Speedtest PNG: "
            f"{error}"
        ) from error

    try:
        image = Image.open(
            BytesIO(response.content)
        ).convert("RGB")

    except Exception as error:
        raise RuntimeError(
            "The Speedtest result was downloaded, but it could "
            "not be opened as an image."
        ) from error

    try:
        ocr_text = pytesseract.image_to_string(
            image,
            config="--psm 6",
        )

    except pytesseract.TesseractNotFoundError as error:
        raise RuntimeError(
            "The Speedtest PNG downloaded successfully, but "
            "Tesseract OCR is not installed or is not available "
            "in PATH on this testing computer."
        ) from error

    except Exception as error:
        raise RuntimeError(
            "TestHub could not read the Speedtest PNG with OCR: "
            f"{error}"
        ) from error

    return parse_speedtest_png_text(
        text=ocr_text,
        result_url=result_url.strip(),
        png_url=png_url,
    )




def parse_speedtest_csv_date(
    value: str,
) -> datetime:
    """
    Parse the Date format exported by the Speedtest desktop app.

    Example:
        Aug/10/2026 16:42
    """
    value = str(value).strip()

    formats = (
        "%b/%d/%Y %H:%M",
        "%b/%d/%Y %I:%M %p",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
    )

    for date_format in formats:
        try:
            return datetime.strptime(
                value,
                date_format,
            )
        except ValueError:
            continue

    raise ValueError(
        f"Unsupported Speedtest Date value: {value}"
    )


SPEEDTEST_CSV_REQUIRED_COLUMNS = {
    "Date",
    "Latency",
    "Download",
    "Upload",
    "ServerName",
}


def find_latest_speedtest_csv_in_downloads() -> Path | None:
    """
    Look in the Windows Downloads folder for the newest Speedtest Result
    History CSV export, instead of relying on a native file-picker dialog
    to find and select it manually - that dialog isn't reliably visible
    on every test PC (it can open off-screen or behind other windows even
    when marked topmost), the same problem the Verizon syslog download
    already works around by reading straight from Downloads.

    Speedtest doesn't use one fixed export filename, so every *.csv in
    Downloads is checked for the columns read_latest_speedtest_csv_results
    requires; the most recently modified one that has them wins.
    """
    downloads_candidates = [
        Path.home() / "Downloads",
    ]

    # OneDrive can redirect Downloads on some test PCs.
    onedrive = os.environ.get("OneDrive")
    if onedrive:
        downloads_candidates.append(
            Path(onedrive) / "Downloads"
        )

    candidates: list[Path] = []

    for downloads_folder in downloads_candidates:
        if not downloads_folder.exists():
            continue

        for path in downloads_folder.glob("*.csv"):
            if not path.is_file():
                continue

            try:
                with path.open(
                    "r",
                    encoding="utf-8-sig",
                    newline="",
                ) as csv_file:
                    fieldnames = set(
                        csv.DictReader(csv_file).fieldnames or []
                    )
            except (OSError, UnicodeDecodeError, csv.Error):
                continue

            if SPEEDTEST_CSV_REQUIRED_COLUMNS.issubset(fieldnames):
                candidates.append(path)

    if not candidates:
        return None

    return max(
        candidates,
        key=lambda path: path.stat().st_mtime,
    )


def read_latest_speedtest_csv_results(
    csv_path: Path,
) -> list[dict[str, Any]]:
    """
    Read the Speedtest Result History CSV and return ALL tests from the
    most recent calendar date in the file.

    Example:
        Aug/10/2026 09:00
        Aug/10/2026 10:15
        Aug/10/2026 12:30
        Aug/10/2026 14:00
        Aug/10/2026 16:42

    If Aug/10/2026 is the newest date, all five rows are returned.
    Older dates are ignored.
    """
    csv_path = Path(csv_path).expanduser().resolve()

    if not csv_path.exists() or not csv_path.is_file():
        raise FileNotFoundError(
            f"Speedtest CSV was not found: {csv_path}"
        )

    parsed_rows: list[tuple[datetime, dict[str, str]]] = []

    with csv_path.open(
        "r",
        encoding="utf-8-sig",
        newline="",
    ) as csv_file:
        reader = csv.DictReader(csv_file)

        fieldnames = reader.fieldnames or []

        missing_columns = sorted(
            SPEEDTEST_CSV_REQUIRED_COLUMNS.difference(fieldnames)
        )

        if missing_columns:
            raise ValueError(
                "The selected file does not look like the expected "
                "Speedtest Result History CSV. Missing columns: "
                + ", ".join(missing_columns)
            )

        for row in reader:
            raw_date = (row.get("Date") or "").strip()

            if not raw_date:
                continue

            try:
                parsed_date = parse_speedtest_csv_date(
                    raw_date
                )
            except ValueError:
                continue

            parsed_rows.append(
                (parsed_date, row)
            )

    if not parsed_rows:
        raise ValueError(
            "No valid Speedtest result rows were found in the CSV."
        )

    newest_calendar_date = max(
        parsed_date.date()
        for parsed_date, _ in parsed_rows
    )

    newest_rows = [
        (parsed_date, row)
        for parsed_date, row in parsed_rows
        if parsed_date.date() == newest_calendar_date
    ]

    newest_rows.sort(
        key=lambda item: item[0]
    )

    results: list[dict[str, Any]] = []

    def optional_float(
        row: dict[str, str],
        key: str,
    ) -> float | None:
        value = (row.get(key) or "").strip()

        if not value:
            return None

        try:
            return float(value)
        except ValueError:
            return None

    for run_number, (parsed_date, row) in enumerate(
        newest_rows,
        start=1,
    ):
        server_name = (
            row.get("ServerName")
            or ""
        ).strip()

        result = {
            "source": "speedtest_result_history_csv",
            "csv_path": str(csv_path),
            "speedtest_id": (
                row.get("Id")
                or ""
            ).strip() or None,
            "timestamp": parsed_date.isoformat(
                timespec="minutes"
            ),
            "run_number": run_number,
            "test_date": parsed_date.isoformat(
                timespec="minutes"
            ),
            "original_date": (
                row.get("Date")
                or ""
            ).strip(),
            "download_mbps": optional_float(
                row,
                "Download",
            ),
            "upload_mbps": optional_float(
                row,
                "Upload",
            ),
            "ping_ms": optional_float(
                row,
                "Latency",
            ),
            "ping_jitter_ms": None,
            "packet_loss_percent": None,
            "server_name": server_name or None,
            "server_location": server_name or None,
            "connection_type": (
                row.get("ConnType")
                or ""
            ).strip() or None,
            "server_latitude": optional_float(
                row,
                "Lat",
            ),
            "server_longitude": optional_float(
                row,
                "Lon",
            ),
            "internal_ip": (
                row.get("InternalIp")
                or ""
            ).strip() or None,
            "external_ip": (
                row.get("ExternalIp")
                or ""
            ).strip() or None,
        }

        results.append(result)

    return results


def write_speedtest_csv_results_workbook(
    workbook_path: Path,
    results: list[dict[str, Any]],
    titan_ip: str,
) -> Path:
    """
    Save every imported result from the newest Speedtest calendar date
    into the same Throughput Results sheet consumed by Analytics.
    """
    workbook_path = Path(workbook_path).resolve()
    workbook_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Throughput Results"

    worksheet.append(ANALYTICS_FIELDS)

    for result in results:
        analytics_result = {
            "timestamp": result.get("timestamp"),
            "run_number": result.get("run_number"),
            "titan_ip": titan_ip,
            "connection_status": None,
            "download_mbps": result.get("download_mbps"),
            "upload_mbps": result.get("upload_mbps"),
            "ping_ms": result.get("ping_ms"),
            "ping_jitter_ms": result.get("ping_jitter_ms"),
            "packet_loss_percent": result.get(
                "packet_loss_percent"
            ),
            "test_duration_seconds": None,
            "isp": None,
            "external_ip": result.get("external_ip"),
            "interface_name": result.get("connection_type"),
            "server_name": result.get("server_name"),
            "server_location": result.get("server_location"),
            "result_url": None,
            "firmware_version": None,
            "carrier": None,
            "technology": None,
            "mode": None,
            "serving_band": None,
            "rsrp_dbm": None,
            "rssi_dbm": None,
            "sinr_db": None,
            "metrics_error": None,
            "notes": (
                "Imported from Speedtest Result History CSV. "
                f"Speedtest ID: {result.get('speedtest_id') or 'N/A'}"
            ),
        }

        worksheet.append([
            analytics_result.get(field)
            for field in ANALYTICS_FIELDS
        ])

    workbook.save(workbook_path)
    workbook.close()

    return workbook_path


def validate_wrapper_session_folder(
    session_folder: Path,
) -> Path:
    session_folder = Path(session_folder).resolve()

    metadata_file = (
        session_folder
        / "metadata"
        / "wrapper_session.json"
    )

    if not metadata_file.exists():
        raise ValueError(
            "Select the wrapper test session root folder, not Downloads or reports."
        )

    (session_folder / "reports").mkdir(
        parents=True,
        exist_ok=True,
    )

    return session_folder


def write_throughput_results_csv(
    csv_path: Path,
    results: list[dict[str, Any]],
) -> Path:
    csv_path = Path(csv_path).resolve()
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = []
    for result in results:
        for key in result.keys():
            if key not in fieldnames:
                fieldnames.append(key)

    with csv_path.open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as csv_file:
        if not fieldnames:
            return csv_path

        writer = csv.DictWriter(
            csv_file,
            fieldnames=fieldnames,
            extrasaction="ignore",
        )
        writer.writeheader()
        writer.writerows(results)

    return csv_path



def write_gui_throughput_workbook(
    workbook_path: Path,
    result: dict[str, Any],
) -> Path:
    """
    Save a GUI-entered throughput result using the same column layout
    consumed by the existing Analytics page.
    """
    workbook_path = Path(workbook_path).resolve()
    workbook_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Throughput Results"

    worksheet.append(ANALYTICS_FIELDS)
    worksheet.append([
        result.get(field)
        for field in ANALYTICS_FIELDS
    ])

    workbook.save(workbook_path)
    workbook.close()

    return workbook_path


def launch_speedtest_gui_job(
    job_id: str,
    request: ThroughputGUILaunchRequest,
) -> None:
    """
    Create a throughput session and launch the Speedtest desktop GUI.

    The engineer chooses/overrides the server directly in Speedtest and
    presses GO manually. TestHub waits for the result to be entered/saved.
    """
    try:
        update_job(
            job_id,
            status="launching",
            message="Opening the Speedtest desktop app.",
            completed_runs=0,
            results=[],
            error=None,
        )

        session_folder = create_session_folder(
            RESULTS_FOLDER,
            "Titan3_GUI",
        )
        create_session_folders(session_folder)

        tester = ThroughputTester(
            mode="gui",
        )
        executable = tester.launch_gui()

        update_job(
            job_id,
            status="waiting_for_result",
            message=(
                "Speedtest is open. Choose the best/default server or "
                "select another server, run the test, then enter the "
                "result in TestHub."
            ),
            completed_runs=0,
            results=[],
            session_folder=str(
                session_folder.resolve()
            ),
            excel_path=None,
            speedtest_gui_executable=str(
                Path(executable).resolve()
            ),
        )

    except Exception as error:
        update_job(
            job_id,
            status="failed",
            message="Could not open the Speedtest desktop app.",
            error=str(error),
        )


def save_speedtest_gui_result(
    request: ThroughputGUIResultRequest,
) -> ThroughputJob:
    """
    Normalize a completed Speedtest GUI result and save it into the same
    workbook format used by Analytics.
    """
    with jobs_lock:
        existing_job = jobs.get(request.job_id)

        if existing_job is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Throughput GUI job was not found.",
            )

        job_snapshot = dict(existing_job)

    existing_results = job_snapshot.get("results") or []
    csv_batch_already_saved = (
        len(existing_results) > 1
        and all(
            result.get("source") == "speedtest_result_history_csv"
            for result in existing_results
        )
    )

    if csv_batch_already_saved:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "This CSV batch is already saved to Analytics. "
                "Do not use Save Result after importing the CSV, because "
                "that would replace the batch with one manual result."
            ),
        )

    session_folder_value = job_snapshot.get(
        "session_folder"
    )

    if not session_folder_value:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "The GUI throughput session folder has not "
                "been created yet."
            ),
        )

    titan = Titan3(
        ip_address=job_snapshot["titan_ip"],
    )

    try:
        reachable = titan.ping()
    except Exception:
        reachable = False

    radio_metrics: dict[str, Any] = {}

    try:
        if reachable:
            metrics = titan.get_radio_metrics()

            if isinstance(metrics, dict):
                radio_metrics = metrics
    except Exception as error:
        radio_metrics = {
            "metrics_error": str(error),
        }

    tester = ThroughputTester(
        mode="gui",
    )

    normalized = tester.normalize_gui_result(
        download_mbps=request.download_mbps,
        upload_mbps=request.upload_mbps,
        ping_ms=request.ping_ms,
        ping_jitter_ms=request.ping_jitter_ms,
        packet_loss_percent=request.packet_loss_percent,
        server_name=request.server_name,
        server_location=request.server_location,
        server_id=request.server_id,
        isp=request.isp,
        result_url=request.result_url,
        notes=request.notes,
    )

    result: dict[str, Any] = {
        "timestamp": datetime.now().isoformat(
            timespec="seconds"
        ),
        "run_number": 1,
        "titan_ip": job_snapshot["titan_ip"],
        "connection_status": (
            "connected"
            if reachable
            else "disconnected"
        ),
        **normalized,
        "firmware_version": radio_metrics.get(
            "firmware_version"
        ),
        "carrier": radio_metrics.get("carrier"),
        "technology": radio_metrics.get(
            "technology"
        ),
        "mode": radio_metrics.get("mode"),
        "serving_band": radio_metrics.get(
            "serving_band"
        ),
        "rsrp_dbm": radio_metrics.get("rsrp_dbm"),
        "rssi_dbm": radio_metrics.get("rssi_dbm"),
        "sinr_db": radio_metrics.get("sinr_db"),
        "metrics_error": radio_metrics.get(
            "metrics_error"
        ),
        "notes": request.notes,
    }

    session_folder = Path(
        session_folder_value
    ).resolve()

    workbook_path = (
        session_folder
        / "reports"
        / "Titan3_GUI_Results.xlsx"
    )

    write_gui_throughput_workbook(
        workbook_path,
        result,
    )

    update_job(
        request.job_id,
        status="completed",
        message=(
            "Speedtest GUI result saved successfully. "
            "It is now available to Analytics."
        ),
        completed_runs=1,
        results=[result],
        excel_path=str(
            workbook_path.resolve()
        ),
        error=None,
    )

    with jobs_lock:
        updated_job = dict(
            jobs[request.job_id]
        )

    return ThroughputJob(**updated_job)


def run_throughput_job(
    job_id: str,
    request: ThroughputRequest,
) -> None:
    try:
        update_job(
            job_id,
            status="running",
            message=(
                f"Running throughput test 1 of "
                f"{request.number_of_runs}."
            ),
            completed_runs=0,
            results=[],
            error=None,
        )

        titan = Titan3(
            ip_address=request.titan_ip,
        )

        session_folder = create_session_folder(
            RESULTS_FOLDER,
            "Titan3_API",
        )

        create_session_folders(
            session_folder
        )

        def report_progress(
            completed_runs: int,
            total_runs: int,
            partial_results: list[dict[str, Any]],
        ) -> None:
            if completed_runs >= total_runs:
                progress_message = (
                    "All throughput test runs are complete."
                )
            else:
                next_run = completed_runs + 1
                progress_message = (
                    f"Completed {completed_runs} of {total_runs} runs. "
                    f"Preparing run {next_run}."
                )

            update_job(
                job_id,
                status="running",
                message=progress_message,
                completed_runs=completed_runs,
                results=partial_results,
            )

        runner = AutomatedTestRunner(
            titan=titan,
            qxdm=None,
            session_folder=session_folder,
            number_of_runs=request.number_of_runs,
            delay_between_runs=request.delay_between_runs,
            timeout_seconds=request.timeout_seconds,
            open_results_after_run=False,
            progress_callback=report_progress,
        )

        results = runner.run()

        update_job(
            job_id,
            status="completed",
            message="Throughput testing completed successfully.",
            completed_runs=len(results),
            results=results,
            session_folder=str(session_folder.resolve()),
            excel_path=str(runner.excel_path.resolve()),
        )

    except Exception as error:
        update_job(
            job_id,
            status="failed",
            message="Throughput testing failed.",
            error=str(error),
        )


# ==========================================================
# QXDM models and state
# ==========================================================

class QXDMStartRequest(BaseModel):
    log_filename: str = Field(
        default=QXDM_DEFAULT_LOG_FILENAME,
        min_length=1,
        max_length=180,
    )
    output_folder: str | None = Field(
        default=None,
        max_length=500,
    )
    max_log_size_mb: int = Field(
        default=QXDM_MAX_LOG_SIZE_MB,
        ge=1,
        le=1024,
    )
    load_mask: bool = True
    continue_without_mask: bool = True
    session_id: str | None = Field(
        default=None,
        max_length=64,
    )
    wrapper_job_id: str | None = Field(
        default=None,
        max_length=100,
    )
    wrapper_session_folder: str | None = Field(
        default=None,
        max_length=500,
    )


class QXDMStatus(BaseModel):
    status: str
    message: str
    workflow_step: str = "idle"
    manual_settings_required: bool = False
    installed: bool
    process_running: bool
    logging_active: bool
    executable_path: str
    mask_path: str | None = None
    expected_log_path: str | None = None
    current_log_path: str | None = None
    current_log_size_bytes: int = 0
    current_log_size_mb: float = 0.0
    current_log_filename: str | None = None
    current_log_modified_at: str | None = None
    max_log_size_mb: int
    started_at: str | None = None
    stopped_at: str | None = None
    session_id: str | None = None
    session_name: str | None = None
    error: str | None = None


qxdm_controller = QXDMController()
qxdm_lock = Lock()

qxdm_state: dict[str, Any] = {
    "status": "idle",
    "message": "QXDM logging is idle.",
    "workflow_step": "idle",
    "manual_settings_required": False,
    "logging_active": False,
    "expected_log_path": None,
    "current_log_path": None,
    "started_at": None,
    "stopped_at": None,
    "session_id": None,
    "session_name": None,
    "error": None,
}


def update_qxdm_state(
    **changes: Any,
) -> None:
    with qxdm_lock:
        qxdm_state.update(changes)


def find_current_qxdm_log() -> Path | None:
    configured_path = qxdm_controller.current_log_path

    if configured_path is None:
        return None

    configured_path = Path(configured_path)

    if configured_path.exists() and configured_path.is_file():
        return configured_path

    try:
        candidates = [
            candidate
            for candidate in configured_path.parent.glob(
                f"{configured_path.stem}*"
            )
            if candidate.is_file()
        ]
    except OSError:
        return None

    if not candidates:
        return None

    try:
        return max(
            candidates,
            key=lambda candidate: candidate.stat().st_mtime,
        )
    except OSError:
        return None


def build_qxdm_status() -> QXDMStatus:
    with qxdm_lock:
        state_snapshot = dict(qxdm_state)

    current_log = find_current_qxdm_log()
    current_size_bytes = 0

    if current_log is not None:
        try:
            current_size_bytes = current_log.stat().st_size
        except OSError:
            current_size_bytes = 0

    configured_mask = qxdm_controller.resolve_default_mask()

    return QXDMStatus(
        status=state_snapshot["status"],
        message=state_snapshot["message"],
        workflow_step=state_snapshot.get("workflow_step", "idle"),
        manual_settings_required=state_snapshot.get(
            "manual_settings_required",
            False,
        ),
        installed=qxdm_controller.executable_exists(),
        process_running=qxdm_controller.is_running(),
        logging_active=state_snapshot["logging_active"],
        executable_path=str(QXDM_EXECUTABLE),
        mask_path=(
            str(configured_mask)
            if configured_mask is not None
            else None
        ),
        expected_log_path=state_snapshot.get(
            "expected_log_path"
        ),
        current_log_path=(
            str(current_log.resolve())
            if current_log is not None
            else None
        ),
        current_log_size_bytes=current_size_bytes,
        current_log_size_mb=round(
            current_size_bytes / (1024 * 1024),
            2,
        ),
        current_log_filename=(
            current_log.name
            if current_log is not None
            else None
        ),
        current_log_modified_at=(
            datetime.fromtimestamp(
                current_log.stat().st_mtime
            ).isoformat(timespec="seconds")
            if current_log is not None
            else None
        ),
        max_log_size_mb=qxdm_controller.max_log_size_mb,
        started_at=state_snapshot["started_at"],
        stopped_at=state_snapshot["stopped_at"],
        session_id=state_snapshot.get("session_id"),
        session_name=state_snapshot.get("session_name"),
        error=state_snapshot["error"],
    )



def resolve_qxdm_session(
    session_id: str | None,
) -> dict[str, Any] | None:
    if not session_id:
        return None

    session = find_session_record(
        results_folder=RESULTS_FOLDER,
        session_id=session_id,
    )

    if session is None:
        raise ValueError(
            "The selected test session was not found."
        )

    return session



def resolve_wrapper_session_folder(
    wrapper_job_id: str | None,
) -> Path | None:
    """Return the wrapper session folder for a valid active wrapper job."""
    if not wrapper_job_id:
        return None

    with wrapper_jobs_lock:
        wrapper_job = wrapper_jobs.get(wrapper_job_id)

        if wrapper_job is None:
            raise ValueError(
                "The selected wrapper session was not found."
            )

        snapshot = dict(wrapper_job)

    result = snapshot.get("result") or {}
    session_folder_value = (
        snapshot.get("session_folder")
        or result.get("session_folder")
    )

    if not session_folder_value:
        raise ValueError(
            "The wrapper session exists but its result folder is not ready yet."
        )

    return Path(session_folder_value).expanduser().resolve()


def write_throughput_results_csv(
    csv_path: Path,
    results: list[dict[str, Any]],
) -> Path:
    """Write every throughput result row to a UTF-8 CSV file."""
    csv_path = Path(csv_path).resolve()
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    fieldnames: list[str] = []

    for result in results:
        for key in result.keys():
            if key not in fieldnames:
                fieldnames.append(key)

    with csv_path.open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as csv_file:
        if not fieldnames:
            return csv_path

        writer = csv.DictWriter(
            csv_file,
            fieldnames=fieldnames,
            extrasaction="ignore",
        )
        writer.writeheader()
        writer.writerows(results)

    return csv_path




def prompt_for_qxdm_wrapper_session_folder(
    initial_path: str | Path | None = None,
) -> Path | None:
    """
    Let the engineer select the wrapper session ROOT folder for QXDM.

    TestHub automatically uses:
        <wrapper session>\qxdm\
    """
    try:
        from tkinter import Tk
        from tkinter.filedialog import askdirectory
    except ImportError as error:
        raise RuntimeError(
            "Tkinter is required to browse for the wrapper session folder."
        ) from error

    initial_dir = None

    if initial_path:
        candidate = Path(initial_path).expanduser()

        if candidate.exists() and candidate.is_dir():
            initial_dir = candidate.resolve()

    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    try:
        selected = askdirectory(
            parent=root,
            title="Select Wrapper Test Session Folder for QXDM",
            initialdir=(
                str(initial_dir)
                if initial_dir is not None
                else None
            ),
            mustexist=True,
        )
    finally:
        root.destroy()

    if not selected:
        return None

    return Path(selected).resolve()


def validate_qxdm_wrapper_session_folder(
    session_folder: Path,
) -> Path:
    session_folder = Path(session_folder).resolve()

    metadata_file = (
        session_folder
        / "metadata"
        / "wrapper_session.json"
    )

    if not metadata_file.exists():
        raise ValueError(
            "Select the wrapper test session root folder, "
            "not the qxdm subfolder or another directory."
        )

    qxdm_folder = session_folder / "qxdm"
    qxdm_folder.mkdir(
        parents=True,
        exist_ok=True,
    )

    return session_folder



def save_qxdm_log_to_session(
    session_id: str | None,
    log_path: Path | None,
    started_at: str | None,
    stopped_at: str | None,
) -> None:
    if not session_id or log_path is None:
        return

    session = find_session_record(
        results_folder=RESULTS_FOLDER,
        session_id=session_id,
    )

    if session is None:
        return

    qxdm_logs = list(
        session.get("qxdm_logs", [])
    )

    resolved_path = str(
        Path(log_path).resolve()
    )

    if not any(
        item.get("log_path") == resolved_path
        for item in qxdm_logs
    ):
        qxdm_logs.append(
            {
                "log_path": resolved_path,
                "filename": Path(resolved_path).name,
                "started_at": started_at,
                "stopped_at": stopped_at,
                "size_bytes": (
                    Path(resolved_path).stat().st_size
                    if Path(resolved_path).exists()
                    else 0
                ),
            }
        )

    session["qxdm_logs"] = qxdm_logs
    session["status"] = "completed"
    session["updated_at"] = datetime.now().isoformat(
        timespec="seconds"
    )

    session_folder = Path(
        session["session_folder"]
    ).resolve()

    metadata_path = (
        session_folder
        / "metadata"
        / "session.json"
    )

    metadata_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    import json

    metadata_path.write_text(
        json.dumps(
            session,
            indent=2,
        ),
        encoding="utf-8",
    )


def run_qxdm_start(
    request: QXDMStartRequest,
) -> None:
    try:
        session = resolve_qxdm_session(
            request.session_id
        )

        wrapper_session_folder = None

        if request.wrapper_session_folder:
            candidate = Path(
                request.wrapper_session_folder
            ).expanduser()

            if candidate.exists() and candidate.is_dir():
                wrapper_session_folder = (
                    validate_qxdm_wrapper_session_folder(
                        candidate
                    )
                )

        # When no wrapper folder was supplied by the frontend, ask the
        # engineer where this QXDM log should belong.
        if wrapper_session_folder is None:
            selected_wrapper = (
                prompt_for_qxdm_wrapper_session_folder(
                    RESULTS_FOLDER
                )
            )

            if selected_wrapper is not None:
                wrapper_session_folder = (
                    validate_qxdm_wrapper_session_folder(
                        selected_wrapper
                    )
                )

        if wrapper_session_folder is not None:
            output_folder = (
                wrapper_session_folder
                / "qxdm"
            )
        else:
            output_folder = (
                Path(request.output_folder).expanduser()
                if request.output_folder
                else RESULTS_FOLDER / "qxdm_logs"
            )

        output_folder = output_folder.resolve()
        output_folder.mkdir(
            parents=True,
            exist_ok=True,
        )

        safe_filename = Path(
            request.log_filename
        ).name.strip()

        if not safe_filename:
            raise ValueError(
                "A QXDM log filename is required."
            )

        if Path(safe_filename).suffix == "":
            safe_filename = f"{safe_filename}.isf"

        timestamp = datetime.now().strftime(
            "%Y%m%d_%H%M%S"
        )

        filename_path = Path(safe_filename)
        timestamped_filename = (
            f"{filename_path.stem}_{timestamp}"
            f"{filename_path.suffix}"
        )

        suggested_log_path = (
            output_folder / timestamped_filename
        ).resolve()

        qxdm_controller.max_log_size_mb = min(
            max(int(request.max_log_size_mb), 1),
            1024,
        )

        session_name = (
            session.get("session_name")
            if session is not None
            else None
        )

        update_qxdm_state(
            status="starting",
            workflow_step="launching",
            message="Launching QXDM and waiting for the diagnostic USB connection.",
            manual_settings_required=False,
            logging_active=False,
            expected_log_path=str(suggested_log_path),
            current_log_path=None,
            started_at=None,
            stopped_at=None,
            session_id=request.session_id,
            session_name=session_name,
            error=None,
        )

        # The controller loads the DMC, opens QXDM Settings, and displays
        # its Continue confirmation window while the user chooses the
        # actual Item Store File save settings.
        update_qxdm_state(
            workflow_step="manual_save_settings",
            message=(
                "QXDM Settings is open. You have 60 seconds to manually "
                "enter the Item Store File values before TestHub continues "
                "automatically."
            ),
            manual_settings_required=True,
        )

        qxdm_controller.start_logging(
            log_path=suggested_log_path,
            load_mask=request.load_mask,
            continue_without_mask=request.continue_without_mask,
        )

        detected_log = find_current_qxdm_log()
        actual_log_path = (
            str(detected_log.resolve())
            if detected_log is not None
            else None
        )

        update_qxdm_state(
            status="logging",
            workflow_step="capture_active",
            message="QXDM capture is active.",
            manual_settings_required=False,
            logging_active=True,
            current_log_path=actual_log_path,
            started_at=datetime.now().isoformat(
                timespec="seconds"
            ),
            stopped_at=None,
            session_id=request.session_id,
            session_name=session_name,
            error=None,
        )

    except Exception as error:
        update_qxdm_state(
            status="failed",
            workflow_step="failed",
            message="QXDM logging failed to start.",
            manual_settings_required=False,
            logging_active=False,
            error=str(error),
        )


def run_qxdm_stop() -> None:
    try:
        with qxdm_lock:
            state_snapshot = dict(qxdm_state)

        update_qxdm_state(
            status="stopping",
            workflow_step="stopping",
            message="Stopping and finalizing the QXDM log.",
            manual_settings_required=False,
            error=None,
        )

        qxdm_controller.stop_logging(
            load_saved_log=True,
        )

        completed_log = find_current_qxdm_log()
        stopped_at = datetime.now().isoformat(
            timespec="seconds"
        )

        save_qxdm_log_to_session(
            session_id=state_snapshot.get("session_id"),
            log_path=completed_log,
            started_at=state_snapshot.get("started_at"),
            stopped_at=stopped_at,
        )

        update_qxdm_state(
            status="completed",
            workflow_step="completed",
            message=(
                "QXDM logging stopped and the log was finalized."
            ),
            manual_settings_required=False,
            logging_active=False,
            current_log_path=(
                str(completed_log.resolve())
                if completed_log is not None
                else None
            ),
            stopped_at=stopped_at,
            error=None,
        )

    except Exception as error:
        update_qxdm_state(
            status="failed",
            workflow_step="failed",
            message="QXDM logging failed to stop cleanly.",
            manual_settings_required=False,
            logging_active=False,
            error=str(error),
        )





# ==========================================================
# Standalone PCAT / ADB RAM dump integration
# ==========================================================

class PCATPathRequest(BaseModel):
    adb_folder: str = Field(
        min_length=1,
        max_length=500,
    )


class PCATLaunchRequest(BaseModel):
    pcat_executable: str | None = Field(
        default=None,
        max_length=500,
    )


pcat_state: dict[str, Any] = {
    "status": "idle",
    "message": "PCAT RAM dump workflow is idle.",
    "adb_folder": None,
    "adb_executable": None,
    "pcat_executable": None,
    "download_mode_result": None,
    "ramdump_result": None,
    "error": None,
}



def get_pcat_config_path() -> Path:
    local_app_data = os.environ.get("LOCALAPPDATA")

    if local_app_data:
        config_dir = Path(local_app_data) / "WNCTestHub"
    else:
        config_dir = Path.home() / ".wnc-testhub"

    config_dir.mkdir(parents=True, exist_ok=True)
    return config_dir / "pcat_config.json"


def load_pcat_config() -> dict[str, Any]:
    config_path = get_pcat_config_path()

    if not config_path.exists():
        return {}

    try:
        data = json.loads(
            config_path.read_text(encoding="utf-8")
        )
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_pcat_config(**changes: Any) -> dict[str, Any]:
    config = load_pcat_config()
    config.update(changes)

    get_pcat_config_path().write_text(
        json.dumps(config, indent=2),
        encoding="utf-8",
    )

    return config


def _valid_adb_candidate(
    candidate: str | Path | None,
) -> Path | None:
    if not candidate:
        return None

    try:
        path = Path(candidate).expanduser().resolve()
    except OSError:
        return None

    if (
        path.exists()
        and path.is_file()
        and path.name.lower() == "adb.exe"
    ):
        return path

    return None


def find_adb_executable() -> tuple[Path | None, str | None]:
    """
    Dynamically locate adb.exe on the current testing computer.

    Search order:
      1. Previously verified path for this PC
      2. ANDROID_SDK_ROOT / ANDROID_HOME
      3. Windows PATH / where.exe
      4. Common Android SDK locations
      5. Limited search of likely user folders
    """
    config = load_pcat_config()

    cached = _valid_adb_candidate(
        config.get("adb_executable")
    )

    if cached is not None:
        return cached, "saved machine setting"

    for env_name in ("ANDROID_SDK_ROOT", "ANDROID_HOME"):
        sdk_root = os.environ.get(env_name)

        if not sdk_root:
            continue

        candidate = _valid_adb_candidate(
            Path(sdk_root) / "platform-tools" / "adb.exe"
        )

        if candidate is not None:
            save_pcat_config(
                adb_executable=str(candidate),
                adb_folder=str(candidate.parent),
            )
            return candidate, env_name

    candidate = _valid_adb_candidate(
        shutil.which("adb")
    )

    if candidate is not None:
        save_pcat_config(
            adb_executable=str(candidate),
            adb_folder=str(candidate.parent),
        )
        return candidate, "Windows PATH"

    if os.name == "nt":
        try:
            completed = subprocess.run(
                ["where.exe", "adb.exe"],
                capture_output=True,
                text=True,
                timeout=5,
                creationflags=subprocess.CREATE_NO_WINDOW,
            )

            if completed.returncode == 0:
                for line in completed.stdout.splitlines():
                    candidate = _valid_adb_candidate(
                        line.strip()
                    )

                    if candidate is not None:
                        save_pcat_config(
                            adb_executable=str(candidate),
                            adb_folder=str(candidate.parent),
                        )
                        return candidate, "where.exe"
        except Exception:
            pass

    local_app_data = os.environ.get("LOCALAPPDATA", "")

    candidates = [
        Path.home()
        / "AppData"
        / "Local"
        / "Android"
        / "Sdk"
        / "platform-tools"
        / "adb.exe",
        (
            Path(local_app_data)
            / "Android"
            / "Sdk"
            / "platform-tools"
            / "adb.exe"
            if local_app_data
            else None
        ),
        Path.home()
        / "Android"
        / "Sdk"
        / "platform-tools"
        / "adb.exe",
        Path("C:/platform-tools/adb.exe"),
        Path("C:/adb/adb.exe"),
        Path("C:/Android/platform-tools/adb.exe"),
    ]

    for raw_candidate in candidates:
        candidate = _valid_adb_candidate(
            raw_candidate
        )

        if candidate is not None:
            save_pcat_config(
                adb_executable=str(candidate),
                adb_folder=str(candidate.parent),
            )
            return candidate, "common platform-tools location"

    search_roots = [
        Path.home() / "Downloads",
        Path.home() / "Documents",
        Path.home() / "Desktop",
        Path.home() / "Projects",
        Path.home() / "Tools",
    ]

    for root in search_roots:
        if not root.exists():
            continue

        try:
            for found in root.rglob("adb.exe"):
                candidate = _valid_adb_candidate(
                    found
                )

                if candidate is None:
                    continue

                parent_parts = {
                    part.lower()
                    for part in candidate.parts[-5:]
                }

                if (
                    "platform-tools" not in parent_parts
                    and "android" not in parent_parts
                    and "adb" not in parent_parts
                ):
                    continue

                save_pcat_config(
                    adb_executable=str(candidate),
                    adb_folder=str(candidate.parent),
                )
                return candidate, f"search under {root}"
        except OSError:
            continue

    return None, None



def open_visible_cmd_with_adb(
    adb_executable: Path,
) -> Path:
    """
    Open a visible Command Prompt in platform-tools and execute the exact
    PCAT preparation commands through a temporary .cmd file.

    Using a .cmd file is more reliable on Windows than passing a long
    command chain directly through cmd.exe /k.
    """
    adb_folder = adb_executable.parent.resolve()

    script_dir = Path(
        os.environ.get(
            "TEMP",
            str(Path.home()),
        )
    ) / "WNCTestHub"

    script_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    script_path = (
        script_dir
        / "pcat_ramdump_setup.cmd"
    )

    script_text = (
        "@echo off\n"
        f'cd /d "{adb_folder}"\n'
        "echo ========================================\n"
        "echo WNC TestHub - PCAT RAM Dump Preparation\n"
        "echo ========================================\n"
        "echo.\n"
        "echo [1/2] Checking download mode...\n"
        "echo.\n"
        "echo adb shell cat /sys/module/qcom_dload_mode/parameters/download_mode\n"
        "adb shell cat /sys/module/qcom_dload_mode/parameters/download_mode\n"
        "echo.\n"
        "echo [2/2] Enabling RAM dump...\n"
        "echo.\n"
        "echo adb shell ./usr/sbin/QMI_VZW_ENABLE_RAMDUMP\n"
        "adb shell ./usr/sbin/QMI_VZW_ENABLE_RAMDUMP\n"
        "echo.\n"
        "echo ========================================\n"
        "echo PCAT preparation commands completed.\n"
        "echo Review the results above, then continue in PCAT.\n"
        "echo ========================================\n"
        "echo.\n"
    )

    script_path.write_text(
        script_text,
        encoding="utf-8",
    )

    subprocess.Popen(
        [
            "cmd.exe",
            "/k",
            str(script_path),
        ],
        cwd=str(adb_folder),
        creationflags=(
            subprocess.CREATE_NEW_CONSOLE
            if os.name == "nt"
            else 0
        ),
    )

    return script_path


def validate_adb_folder(
    adb_folder: str | Path,
) -> tuple[Path, Path]:
    folder = Path(adb_folder).expanduser().resolve()
    adb_executable = folder / "adb.exe"

    if not folder.exists() or not folder.is_dir():
        raise ValueError(
            "The selected platform-tools folder does not exist."
        )

    if not adb_executable.exists() or not adb_executable.is_file():
        raise ValueError(
            "adb.exe was not found in the selected folder. "
            "Select the Android platform-tools folder that contains adb.exe."
        )

    return folder, adb_executable.resolve()


def run_adb_command(
    adb_folder: str | Path,
    shell_arguments: list[str],
    timeout_seconds: int = 30,
) -> dict[str, Any]:
    folder, adb_executable = validate_adb_folder(
        adb_folder
    )

    command = [
        str(adb_executable),
        "shell",
        *shell_arguments,
    ]

    completed = subprocess.run(
        command,
        cwd=str(folder),
        capture_output=True,
        text=True,
        timeout=timeout_seconds,
        creationflags=(
            subprocess.CREATE_NO_WINDOW
            if os.name == "nt"
            else 0
        ),
    )

    stdout = (completed.stdout or "").strip()
    stderr = (completed.stderr or "").strip()

    return {
        "success": completed.returncode == 0,
        "return_code": completed.returncode,
        "stdout": stdout,
        "stderr": stderr,
        "command": " ".join(
            [
                "adb",
                "shell",
                *shell_arguments,
            ]
        ),
        "adb_executable": str(adb_executable),
    }


def prompt_for_adb_folder() -> Path | None:
    try:
        from tkinter import Tk
        from tkinter.filedialog import askdirectory
    except ImportError as error:
        raise RuntimeError(
            "Tkinter is required to browse for the platform-tools folder."
        ) from error

    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    try:
        selected = askdirectory(
            parent=root,
            title="Select Android platform-tools Folder",
            mustexist=True,
        )
    finally:
        root.destroy()

    if not selected:
        return None

    return Path(selected).resolve()


def prompt_for_pcat_executable() -> Path | None:
    try:
        from tkinter import Tk
        from tkinter.filedialog import askopenfilename
    except ImportError as error:
        raise RuntimeError(
            "Tkinter is required to browse for the PCAT executable."
        ) from error

    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    try:
        selected = askopenfilename(
            parent=root,
            title="Select PCAT Executable",
            filetypes=[
                ("Windows applications", "*.exe"),
                ("All files", "*.*"),
            ],
        )
    finally:
        root.destroy()

    if not selected:
        return None

    return Path(selected).resolve()


@app.get("/api/pcat/status")
def get_pcat_status() -> dict[str, Any]:
    config = load_pcat_config()

    cached_adb = _valid_adb_candidate(
        config.get("adb_executable")
    )

    if cached_adb is not None:
        pcat_state["adb_executable"] = str(cached_adb)
        pcat_state["adb_folder"] = str(cached_adb.parent)

    return dict(pcat_state)



@app.get("/api/pcat/find-adb")
def find_pcat_adb() -> dict[str, Any]:
    try:
        adb_executable, discovery_source = find_adb_executable()

        if adb_executable is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=(
                    "TestHub could not find adb.exe automatically. "
                    "Use Browse and select the platform-tools folder manually."
                ),
            )

        adb_folder = adb_executable.parent.resolve()

        pcat_state.update(
            {
                "adb_folder": str(adb_folder),
                "adb_executable": str(adb_executable),
                "status": "adb_configured",
                "message": "ADB was found automatically.",
                "error": None,
            }
        )

        return {
            "success": True,
            "adb_folder": str(adb_folder),
            "adb_executable": str(adb_executable),
            "discovery_source": discovery_source,
            "message": (
                "ADB was found automatically"
                + (
                    f" using {discovery_source}."
                    if discovery_source
                    else "."
                )
            ),
        }

    except HTTPException:
        raise

    except Exception as error:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(error),
        ) from error


@app.post("/api/pcat/open-adb-terminal")
def open_pcat_adb_terminal(
    request: PCATPathRequest,
) -> dict[str, Any]:
    try:
        folder, adb_executable = validate_adb_folder(
            request.adb_folder
        )

        script_path = open_visible_cmd_with_adb(
            adb_executable
        )

        return {
            "success": True,
            "adb_folder": str(folder),
            "adb_executable": str(adb_executable),
            "script_path": str(script_path),
            "message": (
                "Opened Command Prompt and started the two PCAT "
                "ADB setup commands."
            ),
        }

    except ValueError as error:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(error),
        ) from error

    except Exception as error:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(error),
        ) from error


@app.get("/api/pcat/browse-adb")
def browse_pcat_adb_folder() -> dict[str, Any]:
    try:
        selected = prompt_for_adb_folder()

        if selected is None:
            return {
                "cancelled": True,
                "path": None,
            }

        folder, adb_executable = validate_adb_folder(
            selected
        )

        pcat_state.update(
            {
                "adb_folder": str(folder),
                "adb_executable": str(adb_executable),
                "status": "adb_configured",
                "message": "ADB platform-tools folder selected.",
                "error": None,
            }
        )

        save_pcat_config(
            adb_executable=str(adb_executable),
            adb_folder=str(folder),
        )

        return {
            "cancelled": False,
            "path": str(folder),
            "adb_executable": str(adb_executable),
        }

    except ValueError as error:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(error),
        ) from error

    except Exception as error:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(error),
        ) from error


@app.get("/api/pcat/browse-executable")
def browse_pcat_executable() -> dict[str, Any]:
    try:
        selected = prompt_for_pcat_executable()

        if selected is None:
            return {
                "cancelled": True,
                "path": None,
            }

        if (
            not selected.exists()
            or not selected.is_file()
            or selected.suffix.lower() != ".exe"
        ):
            raise ValueError(
                "Select the PCAT Windows executable (.exe)."
            )

        pcat_state.update(
            {
                "pcat_executable": str(selected),
                "message": "PCAT executable selected.",
                "error": None,
            }
        )

        return {
            "cancelled": False,
            "path": str(selected),
        }

    except ValueError as error:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(error),
        ) from error

    except Exception as error:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(error),
        ) from error


@app.post("/api/pcat/check-download-mode")
def pcat_check_download_mode(
    request: PCATPathRequest,
) -> dict[str, Any]:
    try:
        result = run_adb_command(
            request.adb_folder,
            [
                "cat",
                "/sys/module/qcom_dload_mode/parameters/download_mode",
            ],
        )

        if not result["success"]:
            detail = (
                result["stderr"]
                or result["stdout"]
                or "ADB command failed."
            )
            raise RuntimeError(detail)

        pcat_state.update(
            {
                "status": "download_mode_checked",
                "adb_folder": request.adb_folder,
                "adb_executable": result["adb_executable"],
                "download_mode_result": result["stdout"],
                "message": "Download mode checked successfully.",
                "error": None,
            }
        )

        return {
            **result,
            "message": "Download mode checked successfully.",
        }

    except subprocess.TimeoutExpired as error:
        raise HTTPException(
            status_code=status.HTTP_504_GATEWAY_TIMEOUT,
            detail="ADB timed out while checking download mode.",
        ) from error

    except Exception as error:
        pcat_state.update(
            {
                "status": "failed",
                "message": "Could not check download mode.",
                "error": str(error),
            }
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(error),
        ) from error


@app.post("/api/pcat/enable-ramdump")
def pcat_enable_ramdump(
    request: PCATPathRequest,
) -> dict[str, Any]:
    try:
        result = run_adb_command(
            request.adb_folder,
            [
                "./usr/sbin/QMI_VZW_ENABLE_RAMDUMP",
            ],
        )

        if not result["success"]:
            detail = (
                result["stderr"]
                or result["stdout"]
                or "RAM dump command failed."
            )
            raise RuntimeError(detail)

        pcat_state.update(
            {
                "status": "ramdump_enabled",
                "adb_folder": request.adb_folder,
                "adb_executable": result["adb_executable"],
                "ramdump_result": result["stdout"],
                "message": "RAM dump command completed successfully.",
                "error": None,
            }
        )

        return {
            **result,
            "message": "RAM dump command completed successfully.",
        }

    except subprocess.TimeoutExpired as error:
        raise HTTPException(
            status_code=status.HTTP_504_GATEWAY_TIMEOUT,
            detail="ADB timed out while enabling RAM dump.",
        ) from error

    except Exception as error:
        pcat_state.update(
            {
                "status": "failed",
                "message": "Could not enable RAM dump.",
                "error": str(error),
            }
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(error),
        ) from error


@app.post("/api/pcat/open")
def open_pcat_application(
    request: PCATLaunchRequest,
) -> dict[str, Any]:
    try:
        executable_value = (
            request.pcat_executable
            or pcat_state.get("pcat_executable")
        )

        executable: Path | None = None

        if executable_value:
            candidate = Path(
                executable_value
            ).expanduser().resolve()

            if candidate.exists() and candidate.is_file():
                executable = candidate

        if executable is None:
            selected = prompt_for_pcat_executable()

            if selected is None:
                return {
                    "success": False,
                    "cancelled": True,
                    "message": "PCAT executable selection was cancelled.",
                }

            executable = selected

        subprocess.Popen(
            [str(executable)],
            cwd=str(executable.parent),
        )

        pcat_state.update(
            {
                "status": "pcat_open",
                "pcat_executable": str(executable),
                "message": "PCAT opened successfully.",
                "error": None,
            }
        )

        return {
            "success": True,
            "cancelled": False,
            "pcat_executable": str(executable),
            "message": "PCAT opened successfully.",
        }

    except Exception as error:
        pcat_state.update(
            {
                "status": "failed",
                "message": "Could not open PCAT.",
                "error": str(error),
            }
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(error),
        ) from error


# ==========================================================
# Verizon GUI syslog models/state
# ==========================================================

class VerizonSyslogLaunchRequest(BaseModel):
    wrapper_session_folder: str = Field(
        min_length=1,
        max_length=500,
    )
    titan_ip: str = Field(
        default=DEFAULT_TITAN_IP,
        min_length=7,
        max_length=45,
    )


class VerizonSyslogAutomationRequest(BaseModel):
    wrapper_session_folder: str = Field(
        min_length=1,
        max_length=500,
    )
    titan_ip: str = Field(
        default=DEFAULT_TITAN_IP,
        min_length=7,
        max_length=45,
    )
    password: str = Field(
        min_length=1,
        max_length=200,
    )


verizon_syslog_state: dict[str, Any] = {
    "status": "idle",
    "process_running": False,
    "titan_ip": DEFAULT_TITAN_IP,
    "gui_url": None,
    "wrapper_session_folder": None,
    "syslog_folder": None,
    "message": "Verizon GUI syslog workflow is idle.",
    "error": None,
    # Automated login/navigate/download run (see /syslog/verizon/automate).
    # Never put the Verizon GUI password in this dict - it's returned as-is
    # by GET /syslog/verizon/status.
    "automation_status": "idle",
    "automation_step": None,
    "automation_message": None,
    "automation_error": None,
}


def validate_syslog_wrapper_session(
    session_folder: str | Path,
) -> Path:
    session_folder = Path(session_folder).expanduser().resolve()

    metadata_file = (
        session_folder
        / "metadata"
        / "wrapper_session.json"
    )

    if not metadata_file.exists():
        raise ValueError(
            "Select the wrapper test session root folder. "
            "The folder must contain metadata/wrapper_session.json."
        )

    syslog_folder = session_folder / "syslog"
    syslog_folder.mkdir(parents=True, exist_ok=True)

    return session_folder




# ==========================================================
# Unified test wrapper
# ==========================================================

class WrapperStartRequest(BaseModel):
    session_name: str = Field(
        min_length=1,
        max_length=120,
    )
    save_root: str = Field(
        min_length=1,
        max_length=500,
    )
    titan_ip: str = Field(
        default=DEFAULT_TITAN_IP,
        min_length=7,
        max_length=45,
    )

    collect_qxdm: bool = True
    qxdm_mode: str = Field(
        default="manual",
        pattern="^(manual|automatic)$",
    )

    collect_throughput: bool = True
    collect_syslog: bool = True

    number_of_runs: int = Field(
        default=5,
        ge=1,
        le=15,
    )
    delay_between_runs: int = Field(
        default=10,
        ge=0,
        le=3600,
    )
    timeout_seconds: int = Field(
        default=180,
        ge=1,
        le=1800,
    )

    qxdm_log_filename: str = Field(
        default=QXDM_DEFAULT_LOG_FILENAME,
        min_length=1,
        max_length=180,
    )
    qxdm_max_log_size_mb: int = Field(
        default=QXDM_MAX_LOG_SIZE_MB,
        ge=1,
        le=1024,
    )
    load_mask: bool = True
    continue_without_mask: bool = True


class WrapperJob(BaseModel):
    job_id: str
    status: str
    message: str
    qxdm_mode: str
    session_folder: str | None = None
    progress: list[dict[str, Any]] = Field(default_factory=list)
    result: dict[str, Any] | None = None
    error: str | None = None


syslog_controller = SyslogController()
test_wrapper = TestWrapper(
    qxdm=qxdm_controller,
    syslog=syslog_controller,
)

wrapper_jobs: dict[str, dict[str, Any]] = {}
wrapper_jobs_lock = Lock()


def update_wrapper_job(
    job_id: str,
    **changes: Any,
) -> None:
    with wrapper_jobs_lock:
        current = wrapper_jobs.get(job_id)
        if current is None:
            return
        current.update(changes)


def append_wrapper_progress(
    job_id: str,
    item: dict[str, Any],
) -> None:
    with wrapper_jobs_lock:
        current = wrapper_jobs.get(job_id)
        if current is None:
            return
        current.setdefault("progress", []).append(item)
        current["message"] = item.get(
            "message",
            current.get("message", ""),
        )


def run_wrapper_job(
    job_id: str,
    request: WrapperStartRequest,
) -> None:
    """
    Create the wrapper workspace immediately and mark it READY.

    The wrapper is now the shared session/container for the individual
    TestHub tools. Throughput and QXDM are launched from their own pages,
    and those pages use this wrapper job_id to save artifacts back into
    the same wrapper session folder.

    This avoids running a second automatic throughput/QXDM workflow inside
    the wrapper itself and guarantees that pressing Start Test creates the
    session folder right away.
    """
    try:
        update_wrapper_job(
            job_id,
            status="creating",
            message="Creating wrapper session folder.",
            error=None,
        )

        result = test_wrapper.create_workspace(
            save_root=Path(request.save_root),
            session_name=request.session_name,
            titan_ip=request.titan_ip,
            mode="linked",
        )

        # Keep the engineer's requested configuration in wrapper metadata so
        # the session still records what was selected on the Wrapper page.
        result = dict(result)
        result["qxdm_mode"] = request.qxdm_mode
        result["collect_qxdm"] = request.collect_qxdm
        result["collect_throughput"] = request.collect_throughput
        result["collect_syslog"] = request.collect_syslog
        result["number_of_runs"] = request.number_of_runs
        result["delay_between_runs"] = request.delay_between_runs
        result["timeout_seconds"] = request.timeout_seconds
        result["status"] = "ready"

        metadata_path = (
            Path(result["session_folder"])
            / "metadata"
            / "wrapper_session.json"
        )
        test_wrapper.write_metadata(
            metadata_path,
            result,
        )

        update_wrapper_job(
            job_id,
            status="ready",
            message=(
                "Wrapper session created. Open Throughput or QXDM to run "
                "the selected tools; their results will be linked to this "
                "wrapper session automatically."
            ),
            session_folder=result["session_folder"],
            result=result,
            error=None,
        )

    except Exception as error:
        update_wrapper_job(
            job_id,
            status="failed",
            message="Wrapper session could not be created.",
            error=str(error),
        )


# ==========================================================
# Analytics helpers
# ==========================================================

ANALYTICS_FIELDS = [
    "timestamp", "run_number", "titan_ip", "connection_status",
    "download_mbps", "upload_mbps", "ping_ms", "ping_jitter_ms",
    "packet_loss_percent", "test_duration_seconds", "isp",
    "external_ip", "interface_name", "server_name", "server_location",
    "result_url", "firmware_version", "carrier", "technology", "mode",
    "serving_band", "rsrp_dbm", "rssi_dbm", "sinr_db",
    "metrics_error", "notes",
]

def _to_float(value: Any) -> float | None:
    if value is None or value == "": return None
    try: return float(value)
    except (TypeError, ValueError): return None

def load_analytics_history() -> list[dict[str, Any]]:
    history=[]
    if not RESULTS_FOLDER.exists(): return history
    for workbook_path in RESULTS_FOLDER.rglob("*.xlsx"):
        try: wb=load_workbook(workbook_path,read_only=True,data_only=True)
        except Exception: continue
        try:
            if "Throughput Results" not in wb.sheetnames: continue
            for row in wb["Throughput Results"].iter_rows(min_row=2,values_only=True):
                if not any(v is not None for v in row): continue
                rec={f:(row[i] if i < len(row) else None) for i,f in enumerate(ANALYTICS_FIELDS)}
                ts=rec.get("timestamp")
                if isinstance(ts,datetime): rec["timestamp"]=ts.isoformat(timespec="seconds")
                elif ts is not None: rec["timestamp"]=str(ts)
                rec["workbook_path"]=str(workbook_path.resolve())
                rec["session_folder"]=str(workbook_path.parent.parent.resolve())
                history.append(rec)
        finally: wb.close()
    history.sort(key=lambda x:str(x.get("timestamp") or ""),reverse=True)
    return history

def build_analytics_summary(history: list[dict[str, Any]]) -> AnalyticsSummary:
    def vals(k): return [v for x in history if (v:=_to_float(x.get(k))) is not None]
    d,u,p,j=vals("download_mbps"),vals("upload_mbps"),vals("ping_ms"),vals("ping_jitter_ms")
    avg=lambda x: round(sum(x)/len(x),2) if x else None
    return AnalyticsSummary(total_runs=len(history),average_download_mbps=avg(d),minimum_download_mbps=round(min(d),2) if d else None,maximum_download_mbps=round(max(d),2) if d else None,average_upload_mbps=avg(u),minimum_upload_mbps=round(min(u),2) if u else None,maximum_upload_mbps=round(max(u),2) if u else None,average_ping_ms=avg(p),average_jitter_ms=avg(j))


# ==========================================================
# General and device endpoints
# ==========================================================

@app.get("/api/health")
def health_check() -> dict[str, str]:
    return {
        "status": "online",
        "service": "WNC TestHub API",
    }


@app.get("/api/device/status")
def get_device_status(
    titan_ip: str = DEFAULT_TITAN_IP,
) -> dict[str, Any]:
    titan = Titan3(
        ip_address=titan_ip,
    )

    reachable = titan.ping()

    device_status: dict[str, Any] = {
        "ip_address": titan.ip_address,
        "gui_url": titan.gui_url,
        "reachable": reachable,
        "status": (
            "connected"
            if reachable
            else "disconnected"
        ),
        "firmware_version": None,
        "carrier": None,
        "technology": None,
        "mode": None,
        "serving_band": None,
        "rsrp_dbm": None,
        "rssi_dbm": None,
        "sinr_db": None,
        "metrics_error": None,
    }

    if not reachable:
        device_status["metrics_error"] = (
            "Titan is not reachable."
        )

        return device_status

    try:
        metrics = titan.get_radio_metrics()

        if not isinstance(metrics, dict):
            raise TypeError(
                "Titan radio metrics must be returned as a dictionary."
            )

        device_status.update(metrics)

    except Exception as error:
        device_status["metrics_error"] = str(error)

    return device_status





def finalize_wrapper_qxdm_log(
    job_id: str,
) -> dict[str, Any]:
    """
    Let the engineer choose the real QXDM log saved by QXDM, then copy it
    into the wrapper session's qxdm folder.

    This intentionally does not depend on QXDM having used TestHub's expected
    filename or folder. It makes wrapper packaging reliable even when QXDM
    saves under a different filename or directory.
    """
    with wrapper_jobs_lock:
        job = wrapper_jobs.get(job_id)

        if job is None:
            raise ValueError(
                "Wrapper job was not found."
            )

        snapshot = dict(job)

    result = snapshot.get("result") or {}
    session_folder_value = (
        snapshot.get("session_folder")
        or result.get("session_folder")
    )

    if not session_folder_value:
        raise RuntimeError(
            "The wrapper session folder is not available yet."
        )

    selected_log = qxdm_controller.prompt_for_saved_log()

    if selected_log is None:
        raise RuntimeError(
            "No QXDM log was selected."
        )

    selected_log = Path(selected_log).resolve()

    qxdm_folder = (
        Path(session_folder_value).resolve()
        / "qxdm"
    )
    qxdm_folder.mkdir(
        parents=True,
        exist_ok=True,
    )

    destination = (
        qxdm_folder
        / selected_log.name
    ).resolve()

    if selected_log != destination:
        if destination.exists():
            timestamp = datetime.now().strftime(
                "%Y%m%d_%H%M%S"
            )
            destination = (
                qxdm_folder
                / (
                    f"{destination.stem}_{timestamp}"
                    f"{destination.suffix}"
                )
            )

        shutil.copy2(
            selected_log,
            destination,
        )
    else:
        destination = selected_log

    result = dict(result)
    result["qxdm_source_path"] = str(
        selected_log
    )
    result["qxdm_log_path"] = str(
        destination
    )
    result["qxdm_log_filename"] = (
        destination.name
    )
    result["status"] = "completed"

    update_wrapper_job(
        job_id,
        status="completed",
        message=(
            "QXDM log finalized and collected into the "
            "wrapper session folder."
        ),
        session_folder=str(
            Path(session_folder_value).resolve()
        ),
        result=result,
    )

    return {
        "source_path": str(selected_log),
        "saved_path": str(destination),
        "filename": destination.name,
    }



# ==========================================================
# Verizon GUI syslog endpoints
# ==========================================================

@app.get("/api/syslog/verizon/status")
def get_verizon_syslog_status() -> dict[str, Any]:
    return dict(verizon_syslog_state)


@app.get("/api/syslog/verizon/browse-wrapper")
def browse_verizon_syslog_wrapper() -> dict[str, str | None]:
    try:
        from tkinter import Tk
        from tkinter.filedialog import askdirectory

        root = Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        try:
            selected = askdirectory(
                parent=root,
                title="Select WNC TestHub Wrapper Session",
                initialdir=str(Path(RESULTS_FOLDER).resolve()),
                mustexist=True,
            )
        finally:
            root.destroy()

        if not selected:
            return {"path": None}

        session_folder = validate_syslog_wrapper_session(
            selected
        )

        syslog_folder = (
            session_folder / "syslog"
        ).resolve()

        verizon_syslog_state.update(
            {
                "wrapper_session_folder": str(session_folder),
                "syslog_folder": str(syslog_folder),
                "status": "wrapper_selected",
                "message": (
                    "Wrapper session selected. "
                    f"Syslog destination: {syslog_folder}"
                ),
                "error": None,
            }
        )

        return {
            "path": str(session_folder),
            "syslog_folder": str(syslog_folder),
        }

    except ValueError as error:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(error),
        ) from error
    except Exception as error:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(error),
        ) from error


@app.post("/api/syslog/verizon/open")
def open_verizon_gui_for_syslog(
    request: VerizonSyslogLaunchRequest,
) -> dict[str, Any]:
    try:
        wrapper_folder = validate_syslog_wrapper_session(
            request.wrapper_session_folder
        )
        syslog_folder = (
            wrapper_folder / "syslog"
        ).resolve()

        titan_ip = request.titan_ip.strip()
        gui_url = f"https://{titan_ip}/#/login/"

        opened = webbrowser.open(
            gui_url,
            new=2,
        )

        verizon_syslog_state.update(
            {
                "status": "gui_opened",
                "process_running": bool(opened),
                "titan_ip": titan_ip,
                "gui_url": gui_url,
                "wrapper_session_folder": str(wrapper_folder),
                "syslog_folder": str(syslog_folder),
                "message": (
                    "Verizon GUI opened in the default browser. "
                    "Navigate to Diagnostic Monitoring > System Logging, "
                    f"then save the log to: {syslog_folder}"
                ),
                "error": None,
            }
        )

        return dict(verizon_syslog_state)

    except ValueError as error:
        verizon_syslog_state.update(
            {
                "status": "failed",
                "process_running": False,
                "message": "Verizon GUI could not be opened.",
                "error": str(error),
            }
        )
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(error),
        ) from error

    except Exception as error:
        verizon_syslog_state.update(
            {
                "status": "failed",
                "process_running": False,
                "message": "Verizon GUI could not be opened.",
                "error": str(error),
            }
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(error),
        ) from error


@app.post("/api/syslog/verizon/automate")
def automate_verizon_syslog(
    request: VerizonSyslogAutomationRequest,
    background_tasks: BackgroundTasks,
) -> dict[str, Any]:
    """
    Drive the Verizon GUI with a real browser (Playwright) instead of
    requiring a person to click through login > Diagnostic Monitoring >
    System Logging > Save. Runs in the background since it takes a while;
    poll GET /syslog/verizon/status for progress via automation_status /
    automation_step / automation_message / automation_error.
    """
    try:
        wrapper_folder = validate_syslog_wrapper_session(
            request.wrapper_session_folder
        )
    except ValueError as error:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(error),
        ) from error

    syslog_folder = (
        wrapper_folder / "syslog"
    ).resolve()

    titan_ip = request.titan_ip.strip()

    verizon_syslog_state.update(
        {
            "titan_ip": titan_ip,
            "wrapper_session_folder": str(wrapper_folder),
            "syslog_folder": str(syslog_folder),
            "automation_status": "running",
            "automation_step": "launching",
            "automation_message": (
                "Starting the Verizon GUI automation..."
            ),
            "automation_error": None,
        }
    )

    background_tasks.add_task(
        run_verizon_syslog_automation,
        titan_ip,
        request.password,
        syslog_folder,
    )

    return {
        "success": True,
        "message": "Verizon GUI automation started.",
        "syslog_folder": str(syslog_folder),
    }


def run_verizon_syslog_automation(
    titan_ip: str,
    password: str,
    syslog_folder: Path,
) -> None:
    def on_progress(step: str, message: str) -> None:
        verizon_syslog_state.update(
            {
                "automation_step": step,
                "automation_message": message,
            }
        )

    try:
        result = automate_verizon_syslog_download(
            titan_ip=titan_ip,
            password=password,
            destination_folder=syslog_folder,
            progress=on_progress,
        )

        verizon_syslog_state.update(
            {
                "status": "file_detected",
                "automation_status": "completed",
                "automation_step": "completed",
                "automation_message": (
                    f"Automated download saved {result.downloaded_filename}."
                ),
                "automation_error": None,
                "message": (
                    f"Automated download saved {result.downloaded_filename} "
                    f"to {syslog_folder}."
                ),
                "error": None,
            }
        )

    except Exception as error:
        verizon_syslog_state.update(
            {
                "automation_status": "failed",
                "automation_error": str(error),
                "automation_message": str(error),
            }
        )


@app.get("/api/syslog/verizon/open-folder")
def open_verizon_syslog_folder(
    syslog_folder: str | None = None,
) -> dict[str, Any]:
    folder_value = (
        syslog_folder
        or verizon_syslog_state.get("syslog_folder")
    )

    if not folder_value:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="No wrapper syslog folder has been selected yet.",
        )

    folder = Path(folder_value).resolve()
    folder.mkdir(parents=True, exist_ok=True)

    try:
        os.startfile(str(folder))
    except AttributeError as error:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Opening folders is supported on Windows only.",
        ) from error

    return {
        "success": True,
        "path": str(folder),
    }





def list_saved_syslog_files(
    syslog_folder: str | Path,
) -> list[dict[str, Any]]:
    """
    Return Verizon GUI syslog files saved in the wrapper syslog folder.

    Expected filename formats:
        messages_SYS.log
        messages_SYS(1).log
        messages_SYS(2).log
        ...

    Unrelated files in the folder are ignored.
    """
    folder = Path(syslog_folder).expanduser().resolve()
    folder.mkdir(parents=True, exist_ok=True)

    syslog_pattern = re.compile(
        r"^messages_SYS(?:\(\d+\))?\.log$",
        re.IGNORECASE,
    )

    items: list[dict[str, Any]] = []

    for path in folder.iterdir():
        if not path.is_file():
            continue

        if not syslog_pattern.fullmatch(path.name):
            continue

        try:
            stat = path.stat()
        except OSError:
            continue

        items.append(
            {
                "filename": path.name,
                "path": str(path.resolve()),
                "size_bytes": stat.st_size,
                "modified_at": datetime.fromtimestamp(
                    stat.st_mtime
                ).isoformat(timespec="seconds"),
            }
        )

    items.sort(
        key=lambda item: item["modified_at"],
        reverse=True,
    )

    return items



def import_latest_downloaded_syslog(
    destination_folder: str | Path,
) -> Path | None:
    """
    Look in the Windows Downloads folder for Verizon GUI syslogs and copy
    the newest matching file into the selected wrapper's syslog folder.

    Expected filenames:
        messages_SYS.log
        messages_SYS(1).log
        messages_SYS(2).log
        ...
    """
    destination = Path(
        destination_folder
    ).expanduser().resolve()
    destination.mkdir(
        parents=True,
        exist_ok=True,
    )

    downloads_candidates = [
        Path.home() / "Downloads",
    ]

    # OneDrive can redirect Downloads on some test PCs.
    onedrive = os.environ.get("OneDrive")
    if onedrive:
        downloads_candidates.append(
            Path(onedrive) / "Downloads"
        )

    syslog_pattern = re.compile(
        r"^messages_SYS(?:\(\d+\))?\.log$",
        re.IGNORECASE,
    )

    candidates: list[Path] = []

    for downloads_folder in downloads_candidates:
        if not downloads_folder.exists():
            continue

        for path in downloads_folder.iterdir():
            if (
                path.is_file()
                and syslog_pattern.fullmatch(path.name)
            ):
                candidates.append(path)

    if not candidates:
        return None

    newest = max(
        candidates,
        key=lambda path: path.stat().st_mtime,
    )

    target = destination / newest.name

    # If a file with the same name already exists in the wrapper, preserve
    # both instead of overwriting the prior capture.
    if target.exists():
        if (
            target.stat().st_size == newest.stat().st_size
            and int(target.stat().st_mtime)
            == int(newest.stat().st_mtime)
        ):
            return target

        stem = newest.stem
        suffix = newest.suffix
        counter = 1

        while target.exists():
            target = (
                destination
                / f"{stem}_imported_{counter}{suffix}"
            )
            counter += 1

    shutil.copy2(
        newest,
        target,
    )

    return target



@app.get("/api/syslog/verizon/files")
def get_verizon_syslog_files(
    syslog_folder: str | None = None,
) -> dict[str, Any]:
    folder_value = (
        syslog_folder
        or verizon_syslog_state.get("syslog_folder")
    )

    if not folder_value:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Select a wrapper and open the Verizon GUI first.",
        )

    folder = Path(folder_value).resolve()

    imported_file = import_latest_downloaded_syslog(
        folder
    )

    files = list_saved_syslog_files(folder)

    if imported_file is not None:
        message = (
            "Detected the newest Verizon syslog from Downloads and copied "
            f"it into the wrapper: {imported_file}"
        )
    elif files:
        message = (
            f"Detected {len(files)} Verizon syslog file(s) in {folder}."
        )
    else:
        message = (
            "No Verizon syslog was found in the wrapper or Downloads. "
            "Expected a file such as messages_SYS.log or messages_SYS(1).log."
        )

    verizon_syslog_state.update({
        "status": "file_detected" if files else verizon_syslog_state.get("status", "idle"),
        "message": message,
        "error": None,
    })

    return {
        "success": True,
        "syslog_folder": str(folder),
        "count": len(files),
        "latest_file": files[0] if files else None,
        "imported_file": (
            str(imported_file)
            if imported_file is not None
            else None
        ),
        "files": files,
        "message": message,
    }



# ==========================================================
# Wrapper endpoints
# ==========================================================

@app.post(
    "/api/wrapper/start",
    response_model=WrapperJob,
    status_code=status.HTTP_202_ACCEPTED,
)
def start_wrapper_test(
    request: WrapperStartRequest,
    background_tasks: BackgroundTasks,
) -> WrapperJob:
    job_id = str(uuid4())

    job = {
        "job_id": job_id,
        "status": "queued",
        "message": "Wrapper test session queued.",
        "qxdm_mode": request.qxdm_mode,
        "session_folder": None,
        "progress": [],
        "result": None,
        "error": None,
    }

    with wrapper_jobs_lock:
        wrapper_jobs[job_id] = job

    background_tasks.add_task(
        run_wrapper_job,
        job_id,
        request,
    )

    return WrapperJob(**job)


@app.get(
    "/api/wrapper/status/{job_id}",
    response_model=WrapperJob,
)
def get_wrapper_status(
    job_id: str,
) -> WrapperJob:
    with wrapper_jobs_lock:
        job = wrapper_jobs.get(job_id)

        if job is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Wrapper job was not found.",
            )

        snapshot = dict(job)

    return WrapperJob(**snapshot)


@app.post("/api/wrapper/qxdm/finalize/{job_id}")
def finalize_wrapper_qxdm(
    job_id: str,
) -> dict[str, Any]:
    try:
        return finalize_wrapper_qxdm_log(
            job_id
        )
    except ValueError as error:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=str(error),
        ) from error
    except Exception as error:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(error),
        ) from error


@app.get("/api/wrapper/syslog/status")
def get_wrapper_syslog_status() -> dict[str, Any]:
    return syslog_controller.status()


def _find_wrapper_session_metadata_files(
    results_folder: Path,
) -> list[Path]:
    """
    Find every wrapper_session.json under results_folder, at any depth.

    This used to be a one-level-deep glob ("*/metadata/wrapper_session.json"),
    which assumed every session sits directly under RESULTS_FOLDER. Sessions
    created while a stale save location pointed inside an existing session
    (a picker opened inside the results folder, an existing session folder
    right there to accidentally drill into) ended up nested two or more
    levels deep instead - and the one-level glob silently never found them,
    so "latest session" kept resolving to the oldest, outermost one no
    matter how many newer sessions existed. Searching recursively finds
    sessions at any nesting depth, including ones created before the nesting
    itself was fixed at the source in TestWrapper.create_workspace.
    """
    return list(
        results_folder.rglob("metadata/wrapper_session.json")
    )


@app.get("/api/wrapper/latest-session")
def get_latest_wrapper_session() -> dict[str, str | None]:
    """
    Return the most recently modified wrapper test session folder under
    RESULTS_FOLDER (one with metadata/wrapper_session.json), so pages
    like Syslog can default to the newest session automatically instead
    of requiring a manual Browse every time a new session is created
    elsewhere in the app - previously, a page kept sending files to
    whatever session was last manually selected, even after a newer
    session had since been created.
    """
    results_folder = Path(RESULTS_FOLDER).resolve()

    if not results_folder.exists():
        return {"session_folder": None}

    metadata_files = _find_wrapper_session_metadata_files(results_folder)

    if not metadata_files:
        return {"session_folder": None}

    newest_metadata_file = max(
        metadata_files,
        key=lambda path: path.stat().st_mtime,
    )

    session_folder = newest_metadata_file.parent.parent

    return {
        "session_folder": str(session_folder.resolve())
    }


@app.get("/api/wrapper/sessions")
def list_wrapper_sessions() -> dict[str, list[dict[str, Any]]]:
    """
    List every wrapper test session under RESULTS_FOLDER, newest first.

    Lets pages offer an explicit "pick a session" control instead of only
    ever defaulting to whatever is newest - some engineers want a specific
    earlier session, not just the latest one.
    """
    results_folder = Path(RESULTS_FOLDER).resolve()

    if not results_folder.exists():
        return {"sessions": []}

    metadata_files = _find_wrapper_session_metadata_files(results_folder)

    sessions: list[dict[str, Any]] = []

    for metadata_file in metadata_files:
        session_folder = metadata_file.parent.parent

        session_name = session_folder.name
        created_at = None

        try:
            metadata = json.loads(
                metadata_file.read_text(encoding="utf-8")
            )
            session_name = metadata.get("session_name", session_name)
            created_at = metadata.get("created_at")
        except (OSError, ValueError):
            pass

        sessions.append(
            {
                "session_folder": str(session_folder.resolve()),
                "session_name": session_name,
                "created_at": created_at,
                "modified_at": metadata_file.stat().st_mtime,
            }
        )

    sessions.sort(key=lambda item: item["modified_at"], reverse=True)

    return {"sessions": sessions}


@app.get("/api/wrapper/browse-folder")
def browse_wrapper_folder(
    current_path: str | None = None,
) -> dict[str, str | None]:
    """
    Open the native Windows folder picker on the backend/test computer.

    The picker starts in the folder currently shown by the frontend when
    current_path is supplied. If no current path is supplied, it opens in
    TestHub's configured RESULTS_FOLDER instead of defaulting to Downloads.
    """
    try:
        from tkinter import Tk
        from tkinter.filedialog import askdirectory

        initial_folder: Path | None = None

        if current_path:
            candidate = Path(
                current_path
            ).expanduser()

            if candidate.exists() and candidate.is_dir():
                initial_folder = candidate.resolve()

            elif candidate.parent.exists():
                initial_folder = candidate.parent.resolve()

        if initial_folder is None:
            results_candidate = Path(
                RESULTS_FOLDER
            ).expanduser()

            if results_candidate.exists():
                initial_folder = results_candidate.resolve()

            else:
                results_candidate.mkdir(
                    parents=True,
                    exist_ok=True,
                )
                initial_folder = results_candidate.resolve()

        root = Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        try:
            selected = askdirectory(
                parent=root,
                title="Choose WNC TestHub Result Folder",
                initialdir=str(initial_folder),
                mustexist=False,
            )
        finally:
            root.destroy()

        return {
            "path": selected or None,
        }

    except Exception as error:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(error),
        ) from error


# ==========================================================
# Analytics endpoints
# ==========================================================

@app.get("/api/analytics/history", response_model=list[dict[str, Any]])
def get_analytics_history() -> list[dict[str, Any]]:
    return load_analytics_history()

@app.get("/api/analytics/summary", response_model=AnalyticsSummary)
def get_analytics_summary() -> AnalyticsSummary:
    return build_analytics_summary(load_analytics_history())

@app.get("/api/analytics", response_model=AnalyticsResponse)
def get_analytics() -> AnalyticsResponse:
    history=load_analytics_history()
    return AnalyticsResponse(summary=build_analytics_summary(history),history=history)


# ==========================================================
# Session endpoints
# ==========================================================

@app.post(
    "/api/sessions",
    response_model=TestSession,
    status_code=status.HTTP_201_CREATED,
)
def create_test_session(
    request: SessionCreateRequest,
) -> TestSession:
    session = create_session_record(
        results_folder=RESULTS_FOLDER,
        session_name=request.session_name,
        titan_ip=request.titan_ip,
        notes=request.notes,
    )

    return TestSession(**session)


@app.get(
    "/api/sessions",
    response_model=list[TestSession],
)
def get_test_sessions() -> list[TestSession]:
    sessions = list_session_records(
        RESULTS_FOLDER
    )

    return [
        TestSession(**session)
        for session in sessions
    ]


@app.get(
    "/api/sessions/{session_id}",
    response_model=TestSession,
)
def get_test_session(
    session_id: str,
) -> TestSession:
    session = find_session_record(
        results_folder=RESULTS_FOLDER,
        session_id=session_id,
    )

    if session is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Test session was not found.",
        )

    return TestSession(**session)


# ==========================================================
# Throughput endpoints
# ==========================================================

@app.post(
    "/api/throughput/gui/launch",
    response_model=ThroughputJob,
    status_code=status.HTTP_202_ACCEPTED,
)
def launch_throughput_gui(
    request: ThroughputGUILaunchRequest,
    background_tasks: BackgroundTasks,
) -> ThroughputJob:
    job_id = str(uuid4())

    job = {
        "job_id": job_id,
        "status": "queued",
        "message": "Preparing the Speedtest desktop app.",
        "titan_ip": request.titan_ip,
        "number_of_runs": 1,
        "completed_runs": 0,
        "results": [],
        "error": None,
        "session_folder": None,
        "excel_path": None,
    }

    with jobs_lock:
        jobs[job_id] = job

    background_tasks.add_task(
        launch_speedtest_gui_job,
        job_id,
        request,
    )

    return ThroughputJob(**job)


@app.post(
    "/api/throughput/gui/save",
    response_model=ThroughputJob,
)
def save_throughput_gui_result(
    request: ThroughputGUIResultRequest,
) -> ThroughputJob:
    return save_speedtest_gui_result(
        request
    )


@app.post("/api/throughput/gui/import-link")
def import_speedtest_result_link(
    request: SpeedtestResultLinkRequest,
) -> dict[str, Any]:
    """
    Convert the copied Speedtest result URL to its PNG form, download it,
    OCR the result values, and return a preview to the frontend.

    Nothing is saved automatically. The engineer can review/correct the
    imported values before using the existing GUI save endpoint.
    """
    try:
        extracted = fetch_speedtest_png_result(
            request.result_url
        )

    except ValueError as error:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(error),
        ) from error

    except Exception as error:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=str(error),
        ) from error

    required_fields = (
        "download_mbps",
        "upload_mbps",
        "ping_ms",
    )

    detected_fields = [
        field
        for field in required_fields
        if extracted.get(field) is not None
    ]

    success = (
        len(detected_fields)
        == len(required_fields)
    )

    return {
        "success": success,
        "message": (
            "Speedtest PNG imported successfully. Review the "
            "detected values, then save the result."
            if success
            else (
                "The Speedtest PNG was downloaded and read, but "
                "some required values were not detected. Any values "
                "that were found have still been returned so they can "
                "be reviewed or corrected manually."
            )
        ),
        "detected_fields": detected_fields,
        "result": extracted,
    }


@app.post("/api/throughput/gui/import-csv-latest")
def import_latest_speedtest_csv_results(
    request: ThroughputCSVImportRequest,
) -> dict[str, Any]:
    with jobs_lock:
        existing_job = jobs.get(request.job_id)

        if existing_job is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Throughput GUI job was not found.",
            )

        job_snapshot = dict(existing_job)

    try:
        selected_csv = find_latest_speedtest_csv_in_downloads()

        if selected_csv is None:
            return {
                "success": False,
                "cancelled": True,
                "message": (
                    "No Speedtest Result History CSV was found in "
                    "Downloads. In Speedtest, open Result History and "
                    "export it as a CSV, then try again."
                ),
                "results": [],
            }

        imported_results = read_latest_speedtest_csv_results(
            selected_csv
        )

        session_folder_value = job_snapshot.get(
            "session_folder"
        )

        if not session_folder_value:
            raise RuntimeError(
                "The throughput session folder is not available yet."
            )

        throughput_session_folder = Path(
            session_folder_value
        ).resolve()

        analytics_workbook_path = (
            throughput_session_folder
            / "reports"
            / "Titan3_GUI_Results.xlsx"
        )

        write_speedtest_csv_results_workbook(
            workbook_path=analytics_workbook_path,
            results=imported_results,
            titan_ip=job_snapshot["titan_ip"],
        )

        latest_result = imported_results[-1]

        # The CSV is parsed and saved to this job's own Analytics workbook
        # as of here. Update the job now so the page reflects the import
        # immediately - the wrapper folder below is a second, optional
        # destination and must not be able to erase this result if it's
        # cancelled or the chosen folder turns out to be invalid.
        update_job(
            request.job_id,
            status="completed",
            message=(
                f"Imported {len(imported_results)} Speedtest result(s)."
            ),
            completed_runs=len(imported_results),
            number_of_runs=len(imported_results),
            results=imported_results,
            excel_path=str(
                analytics_workbook_path.resolve()
            ),
            error=None,
        )

        response_payload: dict[str, Any] = {
            "success": True,
            "wrapper_saved": False,
            "cancelled": False,
            "message": (
                f"Imported {len(imported_results)} test(s) to Analytics."
            ),
            "latest_date": (
                latest_result.get("original_date", "").split(" ")[0]
            ),
            "count": len(imported_results),
            "latest_result": latest_result,
            "results": imported_results,
            "excel_path": str(
                analytics_workbook_path.resolve()
            ),
            "wrapper_session_folder": None,
            "wrapper_excel_path": None,
            "wrapper_csv_path": None,
        }

        selected_wrapper_folder = (
            Path(request.wrapper_session_folder).expanduser().resolve()
            if request.wrapper_session_folder
            else None
        )

        if selected_wrapper_folder is None:
            response_payload["message"] = (
                f"Imported {len(imported_results)} test(s) to Analytics. "
                "No wrapper session was selected, so nothing was copied "
                "to a wrapper session."
            )

            return response_payload

        try:
            wrapper_session_folder = (
                validate_wrapper_session_folder(
                    selected_wrapper_folder
                )
            )

        except ValueError as error:
            response_payload["message"] = (
                f"Imported {len(imported_results)} test(s) to Analytics, "
                f"but the wrapper session wasn't saved: {error}"
            )

            return response_payload

        wrapper_reports_folder = (
            wrapper_session_folder / "reports"
        ).resolve()

        wrapper_excel_path = (
            wrapper_reports_folder
            / "Titan3_GUI_Results.xlsx"
        )

        wrapper_csv_path = (
            wrapper_reports_folder
            / "throughput_results.csv"
        )

        write_speedtest_csv_results_workbook(
            workbook_path=wrapper_excel_path,
            results=imported_results,
            titan_ip=job_snapshot["titan_ip"],
        )

        write_throughput_results_csv(
            csv_path=wrapper_csv_path,
            results=imported_results,
        )

        response_payload.update({
            "wrapper_saved": True,
            "message": (
                f"Imported and saved {len(imported_results)} test(s) to "
                "Analytics and the wrapper session."
            ),
            "wrapper_session_folder": str(
                wrapper_session_folder
            ),
            "wrapper_excel_path": str(
                wrapper_excel_path.resolve()
            ),
            "wrapper_csv_path": str(
                wrapper_csv_path.resolve()
            ),
        })

        return response_payload

    except ValueError as error:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(error),
        ) from error

    except Exception as error:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(error),
        ) from error


@app.post(
    "/api/throughput/start",
    response_model=ThroughputJob,
    status_code=status.HTTP_202_ACCEPTED,
)
def start_throughput_test(
    request: ThroughputRequest,
    background_tasks: BackgroundTasks,
) -> ThroughputJob:
    job_id = str(uuid4())

    job = {
        "job_id": job_id,
        "status": "queued",
        "message": "Throughput testing has been queued.",
        "titan_ip": request.titan_ip,
        "number_of_runs": request.number_of_runs,
        "completed_runs": 0,
        "results": [],
        "error": None,
        "session_folder": None,
        "excel_path": None,
    }

    with jobs_lock:
        jobs[job_id] = job

    background_tasks.add_task(
        run_throughput_job,
        job_id,
        request,
    )

    return ThroughputJob(**job)


@app.get(
    "/api/throughput/status/{job_id}",
    response_model=ThroughputJob,
)
def get_throughput_status(
    job_id: str,
) -> ThroughputJob:
    with jobs_lock:
        job = jobs.get(job_id)

        if job is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Throughput job was not found.",
            )

        return ThroughputJob(**job)


@app.get(
    "/api/throughput/results/{job_id}",
    response_model=ThroughputJob,
)
def get_throughput_results(
    job_id: str,
) -> ThroughputJob:
    with jobs_lock:
        job = jobs.get(job_id)

        if job is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Throughput job was not found.",
            )

        if job["status"] not in {"completed", "failed"}:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Throughput testing is still running.",
            )

        return ThroughputJob(**job)


# ==========================================================
# QXDM endpoints
# ==========================================================

@app.get(
    "/api/qxdm/status",
    response_model=QXDMStatus,
)
def get_qxdm_status() -> QXDMStatus:
    return build_qxdm_status()


@app.post(
    "/api/qxdm/start",
    response_model=QXDMStatus,
    status_code=status.HTTP_202_ACCEPTED,
)
def start_qxdm_logging(
    request: QXDMStartRequest,
    background_tasks: BackgroundTasks,
) -> QXDMStatus:
    with qxdm_lock:
        if qxdm_state["status"] in {
            "starting",
            "logging",
            "stopping",
        }:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=(
                    "QXDM is already starting, logging, "
                    "or stopping."
                ),
            )

    if not qxdm_controller.executable_exists():
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=(
                "QXDM executable was not found at "
                f"{QXDM_EXECUTABLE}."
            ),
        )

    update_qxdm_state(
        status="starting",
        workflow_step="queued",
        message="QXDM logging start has been queued.",
        manual_settings_required=False,
        logging_active=False,
        error=None,
    )

    background_tasks.add_task(
        run_qxdm_start,
        request,
    )

    return build_qxdm_status()


@app.post(
    "/api/qxdm/saved-log/select",
    response_model=QXDMStatus,
)
def select_saved_qxdm_log() -> QXDMStatus:
    """
    Let the user identify the actual QXDM log saved on this Windows machine.

    This does not change the existing QXDM start/stop workflow. It is a
    fallback for captures whose final filename or folder was changed manually
    inside QXDM Settings.
    """
    try:
        selected_log = qxdm_controller.prompt_for_saved_log()

        if selected_log is None:
            return build_qxdm_status()

        with qxdm_lock:
            state_snapshot = dict(qxdm_state)

        stopped_at = (
            state_snapshot.get("stopped_at")
            or datetime.now().isoformat(timespec="seconds")
        )

        update_qxdm_state(
            current_log_path=str(selected_log.resolve()),
            stopped_at=stopped_at,
            message="Saved QXDM log selected successfully.",
            error=None,
        )

        save_qxdm_log_to_session(
            session_id=state_snapshot.get("session_id"),
            log_path=selected_log,
            started_at=state_snapshot.get("started_at"),
            stopped_at=stopped_at,
        )

        return build_qxdm_status()

    except Exception as error:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(error),
        ) from error


@app.post(
    "/api/qxdm/saved-log/open-folder",
    response_model=QXDMStatus,
)
def open_saved_qxdm_log_folder() -> QXDMStatus:
    """
    Open File Explorer with the currently tracked saved QXDM log selected.
    """
    current_log = find_current_qxdm_log()

    if current_log is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=(
                "No saved QXDM log is currently tracked. "
                "Select the saved log first."
            ),
        )

    try:
        qxdm_controller.open_saved_log_folder(
            current_log
        )
    except Exception as error:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(error),
        ) from error

    return build_qxdm_status()


@app.post(
    "/api/qxdm/stop",
    response_model=QXDMStatus,
    status_code=status.HTTP_202_ACCEPTED,
)
def stop_qxdm_logging(
    background_tasks: BackgroundTasks,
) -> QXDMStatus:
    with qxdm_lock:
        if not qxdm_state["logging_active"]:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="QXDM logging is not currently active.",
            )

        if qxdm_state["status"] == "stopping":
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="QXDM logging is already stopping.",
            )

    update_qxdm_state(
        status="stopping",
        workflow_step="stopping",
        message="QXDM logging stop has been queued.",
        manual_settings_required=False,
        error=None,
    )

    background_tasks.add_task(
        run_qxdm_stop,
    )

    return build_qxdm_status()