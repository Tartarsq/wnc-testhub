from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelResultWriter:
    """Write individual throughput runs and summary statistics to Excel."""

    RESULT_HEADERS = [
        "Timestamp",
        "Run Number",
        "Titan IP",
        "Connection Status",
        "Download Mbps",
        "Upload Mbps",
        "Ping ms",
        "Jitter ms",
        "Packet Loss %",
        "Test Duration Seconds",
        "ISP",
        "External IP",
        "Interface Name",
        "Server Name",
        "Server Location",
        "Result URL",
        "Firmware Version",
        "Carrier",
        "Technology",
        "Mode",
        "Serving Band",
        "RSRP dBm",
        "RSSI dBm",
        "SINR dB",
        "Metrics Error",
        "Notes",
    ]

    RESULT_FIELDS = [
        "timestamp",
        "run_number",
        "titan_ip",
        "connection_status",
        "download_mbps",
        "upload_mbps",
        "ping_ms",
        "ping_jitter_ms",
        "packet_loss_percent",
        "test_duration_seconds",
        "isp",
        "external_ip",
        "interface_name",
        "server_name",
        "server_location",
        "result_url",
        "firmware_version",
        "carrier",
        "technology",
        "mode",
        "serving_band",
        "rsrp_dbm",
        "rssi_dbm",
        "sinr_db",
        "metrics_error",
        "notes",
    ]

    def __init__(self, output_path: Path) -> None:
        self.output_path = Path(output_path).resolve()
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self._ensure_workbook()

    def _ensure_workbook(self) -> None:
        if self.output_path.exists():
            return

        workbook = Workbook()
        results_sheet = workbook.active
        results_sheet.title = "Throughput Results"
        results_sheet.append(self.RESULT_HEADERS)
        self._format_header(results_sheet)
        self._autosize(results_sheet)

        summary_sheet = workbook.create_sheet("Summary")
        summary_sheet.append(["Metric", "Minimum", "Maximum", "Average"])
        self._format_header(summary_sheet)

        workbook.save(self.output_path)

    @staticmethod
    def _format_header(sheet) -> None:
        fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        font = Font(color="FFFFFF", bold=True)

        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    @staticmethod
    def _autosize(sheet) -> None:
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            sheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 12),
                40,
            )

    def append_result(self, result: dict[str, Any]) -> None:
        workbook = load_workbook(self.output_path)
        sheet = workbook["Throughput Results"]

        sheet.append([result.get(field) for field in self.RESULT_FIELDS])
        self._autosize(sheet)
        workbook.save(self.output_path)

    def write_summary(
        self,
        summary: dict[str, dict[str, float | None]],
    ) -> None:
        workbook = load_workbook(self.output_path)

        if "Summary" in workbook.sheetnames:
            sheet = workbook["Summary"]
            workbook.remove(sheet)

        sheet = workbook.create_sheet("Summary")
        sheet.append(["Metric", "Minimum", "Maximum", "Average"])

        for metric, values in summary.items():
            sheet.append([
                metric,
                values.get("minimum"),
                values.get("maximum"),
                values.get("average"),
            ])

        self._format_header(sheet)
        self._autosize(sheet)

        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=4):
            for cell in row:
                cell.number_format = "0.00"

        workbook.save(self.output_path)